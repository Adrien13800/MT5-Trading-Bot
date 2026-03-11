#!/usr/bin/env python3
"""
backtest_optimizer_r3.py - Round 3: maximisation du rendement

Base: H1 2 barres + cooldown 3 (meilleure variante round 1)
Focus: gestion de trade (trailing stop, break-even, SL/TP tuning)

Axes testes:
  1. ATR SL multiplier: 1.0, 1.2, 1.5 (base), 1.8, 2.0, 2.5
  2. R:R ratios: fixes et mixtes (flat/trending)
  3. Break-Even: deplacer SL a l'entree apres X% du TP atteint
  4. Trailing Stop: suivre le prix apres activation
  5. Time exit: fermer apres N bougies si ni SL ni TP touche
  6. Combinaisons des meilleurs
"""

import sys
import os

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import List, Dict, Optional
import pandas as pd
import numpy as np

_root = os.path.dirname(os.path.abspath(__file__))
_backtest_dir = os.path.join(_root, 'backtest')
for p in [_root, _backtest_dir]:
    if p not in sys.path:
        sys.path.insert(0, p)

import MetaTrader5 as mt5
import strategy_core
from strategy_core import TradeType, TradingSession


# ============================================================================
# CONFIG / MT5 / DATA (repris de l'optimizer WR)
# ============================================================================

def load_config():
    try:
        import config
        return {
            'symbols': getattr(config, 'SYMBOLS', ['US30.cash', 'US100.cash', 'US500.cash']),
            'use_daily_preferred_symbol': getattr(config, 'USE_DAILY_PREFERRED_SYMBOL', True),
            'one_symbol_at_a_time': getattr(config, 'ONE_SYMBOL_AT_A_TIME', True),
            'preferred_symbol_by_day': getattr(config, 'PREFERRED_SYMBOL_BY_DAY', None),
            'risk_percent': getattr(config, 'RISK_PERCENT', 1.0),
            'initial_balance': getattr(config, 'INITIAL_BALANCE', 10000.0),
            'login': config.MT5_LOGIN,
            'password': config.MT5_PASSWORD,
            'server': config.MT5_SERVER,
        }
    except Exception as e:
        print(f"ERREUR config: {e}")
        sys.exit(1)


def connect_mt5(cfg):
    import time as _time
    mt5.shutdown()  # clean state
    _time.sleep(1)
    paths = [r"C:\Program Files\MetaTrader\terminal64.exe",
             r"C:\Program Files\MetaTrader 5\terminal64.exe",
             r"C:\Program Files\MetaTrader 5 Risk Manager\terminal64.exe"]
    ok = mt5.initialize()
    if not ok:
        for p in paths:
            if os.path.exists(p):
                ok = mt5.initialize(path=p)
                if ok:
                    break
    if not ok:
        # Last resort: wait for terminal to start
        for p in paths:
            if os.path.exists(p):
                os.startfile(p)
                print(f"Lancement de MT5: {p}")
                for _ in range(30):
                    _time.sleep(2)
                    if mt5.initialize(path=p):
                        ok = True
                        break
                if ok:
                    break
    if not ok:
        print(f"Impossible d'initialiser MT5. Erreur: {mt5.last_error()}")
        print("Assurez-vous que MetaTrader 5 est ouvert.")
        sys.exit(1)
    if not mt5.login(login=cfg['login'], password=cfg['password'], server=cfg['server']):
        print(f"Login echoue: {mt5.last_error()}")
        sys.exit(1)
    info = mt5.account_info()
    print(f"Connecte: {info.login} @ {info.server}")


def load_data(symbols):
    m5_data, h1_data = {}, {}
    for sym in symbols:
        si = mt5.symbol_info(sym)
        if si is None:
            continue
        if not si.visible:
            mt5.symbol_select(sym, True)
        best = None
        for n in [50000, 100000, 200000, 500000, 1000000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_M5, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            strategy_core.compute_indicators(df)
            m5_data[sym] = df
            print(f"  {sym}: {len(df)} M5")
        best = None
        for n in [10000, 50000, 100000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_H1, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            h1_data[sym] = df
    return m5_data, h1_data


# Cache symbol_info to avoid thousands of MT5 API calls per variant
_symbol_info_cache = {}

def get_symbol_info(symbol):
    if symbol not in _symbol_info_cache:
        si = mt5.symbol_info(symbol)
        if si:
            _symbol_info_cache[symbol] = si
    return _symbol_info_cache.get(symbol)


def calc_lot(symbol, entry, sl, balance, risk_pct):
    si = get_symbol_info(symbol)
    if not si:
        return 0.0
    sd = abs(entry - sl)
    if sd <= 0:
        return 0.0
    risk = balance * (risk_pct / 100.0)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        rpl = (sd / ts) * tv
    elif si.trade_contract_size > 0:
        rpl = (sd * si.trade_contract_size) / entry
    else:
        rpl = sd * si.point
    if rpl <= 0:
        return 0.0
    lot = risk / rpl
    lot = max(si.volume_min, min(lot, si.volume_max))
    if si.volume_step > 0:
        lot = (lot // si.volume_step) * si.volume_step
    return round(lot, 2)


def calc_profit(symbol, entry, exit_p, lot, ttype):
    si = get_symbol_info(symbol)
    if not si:
        return 0.0
    diff = (exit_p - entry) if ttype == TradeType.LONG else (entry - exit_p)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        return (diff / ts) * tv * lot
    if si.trade_contract_size > 0:
        return (diff * si.trade_contract_size * lot) / entry
    return diff * si.point * lot


# ============================================================================
# SIMULATION Round 3 - avec gestion avancee des trades
# ============================================================================

@dataclass
class SimTrade:
    symbol: str
    trade_type: TradeType
    entry_time: datetime
    entry_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    original_sl: float = 0.0       # SL initial (avant BE/trailing)
    be_triggered: bool = False      # break-even active
    trail_active: bool = False      # trailing actif
    bars_open: int = 0              # nombre de bougies depuis ouverture
    max_favorable: float = 0.0     # prix max favorable atteint
    exit_time: Optional[datetime] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    profit: float = 0.0


def run_variant(m5_data, h1_data, cfg,
                # SL / TP params
                atr_sl_mult=1.5,
                rr_flat=1.0, rr_trending=1.5,
                # Break-Even: deplacer SL a entry apres que le prix atteigne X% du TP
                be_trigger_pct=0.0,      # 0 = desactive, 0.5 = a 50% du TP
                # Trailing Stop: trail le SL a distance ATR apres activation
                trail_atr_mult=0.0,      # 0 = desactive, 1.0 = trail a 1x ATR
                trail_activate_pct=0.0,  # activer le trail apres X% du TP atteint (0 = apres BE)
                # Time exit: fermer apres N bougies M5 si ni SL ni TP touche
                max_trade_bars=0,        # 0 = desactive
                # Cooldown
                cooldown_bars=3):

    symbols = cfg['symbols']
    use_daily_pref = cfg.get('use_daily_preferred_symbol', True)
    one_at_a_time = cfg.get('one_symbol_at_a_time', True)
    pref_by_day = cfg.get('preferred_symbol_by_day') or {}
    risk_pct = cfg.get('risk_percent', 1.0)
    balance = cfg.get('initial_balance', 10000.0)

    # Override strategy_core params temporarily
    orig_rr_f = strategy_core.RISK_REWARD_RATIO_FLAT
    orig_rr_t = strategy_core.RISK_REWARD_RATIO_TRENDING
    orig_atr = strategy_core.ATR_SL_MULTIPLIER
    strategy_core.RISK_REWARD_RATIO_FLAT = rr_flat
    strategy_core.RISK_REWARD_RATIO_TRENDING = rr_trending
    strategy_core.ATR_SL_MULTIPLIER = atr_sl_mult

    # Build event timeline
    events = []
    for sym in symbols:
        if sym not in m5_data:
            continue
        d = m5_data[sym]
        for i in range(50, len(d)):
            ts = d.index[i]
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            events.append((ts, sym, i))
    events.sort(key=lambda x: x[0])

    open_trades: Dict[str, List[SimTrade]] = {}
    closed_trades: List[SimTrade] = []
    last_bar_time: Dict[str, datetime] = {}
    last_loss_time = None

    for current_bar_time, symbol, bar_index in events:
        df = m5_data[symbol]
        market_data = df.iloc[:bar_index + 1]
        current_bar = df.iloc[bar_index]

        # ============================================================
        # TRADE MANAGEMENT: SL/TP check + BE + Trailing + Time exit
        # ============================================================
        if symbol in open_trades and open_trades[symbol]:
            to_close = []
            for idx, trade in enumerate(open_trades[symbol]):
                trade.bars_open += 1
                hit = None

                # Get ATR for trailing (use current bar's ATR)
                current_atr = market_data['ATR'].iloc[-1] if 'ATR' in market_data.columns else 0

                if trade.trade_type == TradeType.LONG:
                    # Track max favorable price
                    if current_bar['high'] > trade.max_favorable:
                        trade.max_favorable = current_bar['high']

                    tp_distance = trade.take_profit - trade.entry_price
                    current_move = trade.max_favorable - trade.entry_price

                    # Break-Even check
                    if be_trigger_pct > 0 and not trade.be_triggered and tp_distance > 0:
                        if current_move >= tp_distance * be_trigger_pct:
                            trade.stop_loss = trade.entry_price
                            trade.be_triggered = True

                    # Trailing check
                    if trail_atr_mult > 0 and current_atr > 0:
                        activate_dist = tp_distance * trail_activate_pct if trail_activate_pct > 0 else (tp_distance * be_trigger_pct if be_trigger_pct > 0 else tp_distance * 0.5)
                        if current_move >= activate_dist:
                            trade.trail_active = True
                        if trade.trail_active:
                            new_sl = trade.max_favorable - (current_atr * trail_atr_mult)
                            if new_sl > trade.stop_loss:
                                trade.stop_loss = new_sl

                    # SL/TP check
                    if current_bar['low'] <= trade.stop_loss:
                        hit = ('SL' if trade.stop_loss <= trade.entry_price else 'TSL', trade.stop_loss)
                    elif current_bar['high'] >= trade.take_profit:
                        hit = ('TP', trade.take_profit)

                else:  # SHORT
                    # Track max favorable price (lowest)
                    if trade.max_favorable == 0 or current_bar['low'] < trade.max_favorable:
                        trade.max_favorable = current_bar['low']

                    tp_distance = trade.entry_price - trade.take_profit
                    current_move = trade.entry_price - trade.max_favorable

                    # Break-Even check
                    if be_trigger_pct > 0 and not trade.be_triggered and tp_distance > 0:
                        if current_move >= tp_distance * be_trigger_pct:
                            trade.stop_loss = trade.entry_price
                            trade.be_triggered = True

                    # Trailing check
                    if trail_atr_mult > 0 and current_atr > 0:
                        activate_dist = tp_distance * trail_activate_pct if trail_activate_pct > 0 else (tp_distance * be_trigger_pct if be_trigger_pct > 0 else tp_distance * 0.5)
                        if current_move >= activate_dist:
                            trade.trail_active = True
                        if trade.trail_active:
                            new_sl = trade.max_favorable + (current_atr * trail_atr_mult)
                            if new_sl < trade.stop_loss:
                                trade.stop_loss = new_sl

                    # SL/TP check
                    if current_bar['high'] >= trade.stop_loss:
                        hit = ('SL' if trade.stop_loss >= trade.entry_price else 'TSL', trade.stop_loss)
                    elif current_bar['low'] <= trade.take_profit:
                        hit = ('TP', trade.take_profit)

                # Time-based exit
                if hit is None and max_trade_bars > 0 and trade.bars_open >= max_trade_bars:
                    hit = ('TIME', float(current_bar['close']))

                if hit:
                    trade.exit_reason, trade.exit_price = hit
                    trade.exit_time = df.index[bar_index]
                    trade.profit = calc_profit(symbol, trade.entry_price, trade.exit_price,
                                               trade.lot_size, trade.trade_type)
                    balance += trade.profit
                    if trade.profit < 0:
                        last_loss_time = current_bar_time
                    closed_trades.append(trade)
                    to_close.append(idx)
            for idx in reversed(to_close):
                open_trades[symbol].pop(idx)
            if symbol in open_trades and not open_trades[symbol]:
                del open_trades[symbol]

        # ============================================================
        # SIGNAL DETECTION (same as base strategy)
        # ============================================================
        if symbol in last_bar_time and current_bar_time <= last_bar_time[symbol]:
            continue
        last_bar_time[symbol] = current_bar_time

        if use_daily_pref and pref_by_day:
            wd = current_bar_time.weekday() if hasattr(current_bar_time, 'weekday') \
                else current_bar_time.to_pydatetime().weekday()
            if pref_by_day.get(wd) is not None and symbol != pref_by_day.get(wd):
                continue

        # Cooldown
        if cooldown_bars > 0 and last_loss_time is not None:
            if (current_bar_time - last_loss_time).total_seconds() / 300 < cooldown_bars:
                continue

        def has_other_open():
            if not one_at_a_time:
                return False
            for s, tl in open_trades.items():
                if s != symbol and tl:
                    return True
            return False

        # H1 data
        df_h1_f = strategy_core.get_h1_data_at_time(
            h1_data.get(symbol, pd.DataFrame()), current_bar_time
        ) if symbol in h1_data else None

        for allow, check_fn, ttype, sl_fn, tp_sign in [
            (strategy_core.ALLOW_LONG, strategy_core.check_long_signal,
             TradeType.LONG, strategy_core.calculate_sl_long, 1),
            (strategy_core.ALLOW_SHORT, strategy_core.check_short_signal,
             TradeType.SHORT, strategy_core.calculate_sl_short, -1),
        ]:
            if not allow:
                continue

            signal = check_fn(market_data, df_h1_f, symbol)
            if not signal:
                continue

            # Entry price = next bar open
            if bar_index + 1 < len(df):
                entry_price = float(df.iloc[bar_index + 1]['open'])
            else:
                entry_price = float(current_bar['close'])

            stop_loss = sl_fn(market_data, 10)

            if ttype == TradeType.LONG:
                stop_dist = entry_price - stop_loss
                if stop_loss <= 0 or stop_loss >= entry_price:
                    continue
            else:
                stop_dist = stop_loss - entry_price
                if stop_loss <= 0 or stop_loss <= entry_price:
                    continue

            sl_pct = abs(entry_price - stop_loss) / entry_price if entry_price > 0 else 0
            if sl_pct > 0.05:
                continue

            rr = strategy_core.get_risk_reward_ratio(market_data)
            take_profit = entry_price + (stop_dist * rr * tp_sign)
            lot = calc_lot(symbol, entry_price, stop_loss, balance, risk_pct)

            if lot <= 0 or has_other_open():
                continue

            trade = SimTrade(
                symbol=symbol, trade_type=ttype,
                entry_time=current_bar_time,
                entry_price=entry_price, stop_loss=stop_loss,
                take_profit=take_profit, lot_size=lot,
                original_sl=stop_loss,
                max_favorable=entry_price if ttype == TradeType.LONG else entry_price,
            )
            open_trades.setdefault(symbol, []).append(trade)

    # Restore strategy_core params
    strategy_core.RISK_REWARD_RATIO_FLAT = orig_rr_f
    strategy_core.RISK_REWARD_RATIO_TRENDING = orig_rr_t
    strategy_core.ATR_SL_MULTIPLIER = orig_atr

    return closed_trades, balance


def compute_stats(trades, initial, final):
    closed = [t for t in trades if t.exit_reason in ('SL', 'TP', 'TSL', 'TIME')]
    total = len(closed)
    if total == 0:
        return {'trades': 0, 'wr': 0, 'net': 0, 'pf': 0, 'rendement': 0,
                'avg_win': 0, 'avg_loss': 0, 'max_dd': 0, 'be_count': 0, 'tsl_count': 0}
    wins = [t for t in closed if t.profit > 0]
    losses = [t for t in closed if t.profit < 0]
    be_trades = [t for t in closed if abs(t.profit) < 0.5]  # ~break-even
    tsl_trades = [t for t in closed if t.exit_reason == 'TSL']
    tw = sum(t.profit for t in wins)
    tl = abs(sum(t.profit for t in losses))
    equity, peak, max_dd = initial, initial, 0
    for t in sorted(closed, key=lambda x: x.entry_time):
        equity += t.profit
        if equity > peak:
            peak = equity
        dd = peak - equity
        if dd > max_dd:
            max_dd = dd
    return {
        'trades': total, 'wins': len(wins), 'losses': len(losses),
        'wr': len(wins) / total * 100, 'net': final - initial,
        'pf': tw / tl if tl > 0 else 999,
        'rendement': (final - initial) / initial * 100,
        'avg_win': tw / len(wins) if wins else 0,
        'avg_loss': tl / len(losses) if losses else 0,
        'max_dd': max_dd,
        'be_count': len(be_trades),
        'tsl_count': len(tsl_trades),
    }


def main():
    cfg = load_config()
    print("=" * 70)
    print("OPTIMISEUR R3b - BE / Trailing / Time Exit / Combos (cache MT5)")
    print("=" * 70)
    connect_mt5(cfg)
    print("\nChargement des donnees...")
    m5_data, h1_data = load_data(cfg['symbols'])
    initial = cfg.get('initial_balance', 10000.0)

    variants = [
        # === REF: meilleur du R3a ===
        ("REF: RR fixe 1:1.5 (best R3a)",  {"rr_flat": 1.5, "rr_trending": 1.5}),

        # === CAT 1: Break-Even (base RR adaptatif) ===
        ("BE 33% du TP",             {"be_trigger_pct": 0.33}),
        ("BE 50% du TP",             {"be_trigger_pct": 0.50}),
        ("BE 66% du TP",             {"be_trigger_pct": 0.66}),
        ("BE 75% du TP",             {"be_trigger_pct": 0.75}),

        # === CAT 2: BE + meilleur RR (1:1.5 fixe) ===
        ("RR 1:1.5 + BE 33%",       {"rr_flat": 1.5, "rr_trending": 1.5, "be_trigger_pct": 0.33}),
        ("RR 1:1.5 + BE 50%",       {"rr_flat": 1.5, "rr_trending": 1.5, "be_trigger_pct": 0.50}),
        ("RR 1:1.5 + BE 66%",       {"rr_flat": 1.5, "rr_trending": 1.5, "be_trigger_pct": 0.66}),
        ("RR 1:2 + BE 33%",         {"rr_flat": 2.0, "rr_trending": 2.0, "be_trigger_pct": 0.33}),
        ("RR 1:2 + BE 50%",         {"rr_flat": 2.0, "rr_trending": 2.0, "be_trigger_pct": 0.50}),
        ("RR 1:2.5 + BE 33%",       {"rr_flat": 2.5, "rr_trending": 2.5, "be_trigger_pct": 0.33}),
        ("RR 1:3 + BE 33%",         {"rr_flat": 3.0, "rr_trending": 3.0, "be_trigger_pct": 0.33}),

        # === CAT 3: Trailing Stop ===
        ("Trail 1.0x ATR (act 50%)",  {"be_trigger_pct": 0.5, "trail_atr_mult": 1.0, "trail_activate_pct": 0.5}),
        ("Trail 1.5x ATR (act 50%)",  {"be_trigger_pct": 0.5, "trail_atr_mult": 1.5, "trail_activate_pct": 0.5}),
        ("Trail 2.0x ATR (act 50%)",  {"be_trigger_pct": 0.5, "trail_atr_mult": 2.0, "trail_activate_pct": 0.5}),
        ("Trail 1.0x ATR (act 33%)",  {"be_trigger_pct": 0.33, "trail_atr_mult": 1.0, "trail_activate_pct": 0.33}),
        ("Trail 1.5x ATR (act 33%)",  {"be_trigger_pct": 0.33, "trail_atr_mult": 1.5, "trail_activate_pct": 0.33}),

        # === CAT 4: Trail + meilleur RR ===
        ("RR 1:1.5 + Trail 1.0x (act50%)",  {"rr_flat": 1.5, "rr_trending": 1.5,
                                               "be_trigger_pct": 0.5, "trail_atr_mult": 1.0, "trail_activate_pct": 0.5}),
        ("RR 1:1.5 + Trail 1.5x (act50%)",  {"rr_flat": 1.5, "rr_trending": 1.5,
                                               "be_trigger_pct": 0.5, "trail_atr_mult": 1.5, "trail_activate_pct": 0.5}),
        ("RR 1:2 + Trail 1.5x (act50%)",    {"rr_flat": 2.0, "rr_trending": 2.0,
                                               "be_trigger_pct": 0.5, "trail_atr_mult": 1.5, "trail_activate_pct": 0.5}),
        ("RR 1:3 + Trail 1.5x (act33%)",    {"rr_flat": 3.0, "rr_trending": 3.0,
                                               "be_trigger_pct": 0.33, "trail_atr_mult": 1.5, "trail_activate_pct": 0.33}),

        # === CAT 5: Time Exit ===
        ("Time exit 24 bars (2h)",    {"max_trade_bars": 24}),
        ("Time exit 48 bars (4h)",    {"max_trade_bars": 48}),
        ("Time exit 72 bars (6h)",    {"max_trade_bars": 72}),
        ("Time exit 144 bars (12h)",  {"max_trade_bars": 144}),
        ("RR 1:1.5 + Time 48 bars",  {"rr_flat": 1.5, "rr_trending": 1.5, "max_trade_bars": 48}),
        ("RR 1:1.5 + Time 72 bars",  {"rr_flat": 1.5, "rr_trending": 1.5, "max_trade_bars": 72}),
        ("RR 1:2 + Time 48 bars",    {"rr_flat": 2.0, "rr_trending": 2.0, "max_trade_bars": 48}),

        # === CAT 6: Full Combos ===
        ("SL 1.2x + RR 1:1.5",      {"atr_sl_mult": 1.2, "rr_flat": 1.5, "rr_trending": 1.5}),
        ("SL 1.2x + RR 1:2",        {"atr_sl_mult": 1.2, "rr_flat": 2.0, "rr_trending": 2.0}),
        ("RR 1:1.5 + BE 50% + Trail 1.5x", {"rr_flat": 1.5, "rr_trending": 1.5,
                                              "be_trigger_pct": 0.5, "trail_atr_mult": 1.5, "trail_activate_pct": 0.5}),
        ("RR 1:2 + BE 33% + Trail 1.0x",   {"rr_flat": 2.0, "rr_trending": 2.0,
                                              "be_trigger_pct": 0.33, "trail_atr_mult": 1.0, "trail_activate_pct": 0.33}),
        ("SL 1.2x + RR 1:1.5 + BE 50%",   {"atr_sl_mult": 1.2, "rr_flat": 1.5, "rr_trending": 1.5,
                                              "be_trigger_pct": 0.5}),
        ("SL 1.2x + RR 1:2 + BE 50% + Trail 1.5x", {"atr_sl_mult": 1.2, "rr_flat": 2.0, "rr_trending": 2.0,
                                                        "be_trigger_pct": 0.5, "trail_atr_mult": 1.5, "trail_activate_pct": 0.5}),
    ]

    results = []
    n_variants = len(variants)
    print(f"\nTest de {n_variants} variantes...\n")

    for i, (name, kwargs) in enumerate(variants, 1):
        print(f"  [{i:2d}/{n_variants}] {name}...", end=" ", flush=True)
        trades, final = run_variant(m5_data, h1_data, cfg, **kwargs)
        stats = compute_stats(trades, initial, final)
        stats['name'] = name
        results.append(stats)
        print(f"{stats['trades']} trades | WR {stats['wr']:.1f}% | PF {stats['pf']:.2f} | "
              f"Rdt {stats['rendement']:.1f}% | DD {stats['max_dd']:.0f}")

    # === CLASSEMENT PAR RENDEMENT ===
    results_rdt = sorted(results, key=lambda x: x['rendement'], reverse=True)

    print(f"\n{'=' * 130}")
    print(f"{'CLASSEMENT PAR RENDEMENT (TOP 20)':^130}")
    print(f"{'=' * 130}")
    print(f"{'#':>3} {'Variante':<48} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8} {'AvgW':>7} {'AvgL':>7} {'TSL':>4}")
    print("-" * 130)
    for i, s in enumerate(results_rdt[:20], 1):
        marker = " <-- BEST" if i == 1 else ""
        print(f"{i:3d} {s['name']:<48} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$ "
              f"{s.get('avg_win', 0):6.0f}$ {s.get('avg_loss', 0):6.0f}$ {s.get('tsl_count', 0):4d}{marker}")

    # === CLASSEMENT PAR WR ===
    results_wr = sorted(results, key=lambda x: (x['wr'], x['rendement']), reverse=True)

    print(f"\n{'=' * 130}")
    print(f"{'CLASSEMENT PAR WIN RATE (TOP 20)':^130}")
    print(f"{'=' * 130}")
    print(f"{'#':>3} {'Variante':<48} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8}")
    print("-" * 130)
    for i, s in enumerate(results_wr[:20], 1):
        print(f"{i:3d} {s['name']:<48} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$")

    # === CLASSEMENT PAR PF ===
    results_pf = sorted(results, key=lambda x: x['pf'], reverse=True)

    print(f"\n{'=' * 130}")
    print(f"{'CLASSEMENT PAR PROFIT FACTOR (TOP 10)':^130}")
    print(f"{'=' * 130}")
    print(f"{'#':>3} {'Variante':<48} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8}")
    print("-" * 130)
    for i, s in enumerate(results_pf[:10], 1):
        print(f"{i:3d} {s['name']:<48} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$")

    # === CLASSEMENT PAR RENDEMENT/DD (ratio Calmar) ===
    for s in results:
        s['calmar'] = s['rendement'] / s['max_dd'] * initial if s['max_dd'] > 0 else 0
    results_calmar = sorted(results, key=lambda x: x['calmar'], reverse=True)

    print(f"\n{'=' * 130}")
    print(f"{'CLASSEMENT PAR RENDEMENT/DRAWDOWN (TOP 10)':^130}")
    print(f"{'=' * 130}")
    print(f"{'#':>3} {'Variante':<48} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Rdt%':>8} {'MaxDD$':>8} {'Rdt/DD':>8}")
    print("-" * 130)
    for i, s in enumerate(results_calmar[:10], 1):
        ratio = s['rendement'] / (s['max_dd'] / initial * 100) if s['max_dd'] > 0 else 0
        print(f"{i:3d} {s['name']:<48} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['rendement']:7.1f}% {s['max_dd']:7.0f}$ {ratio:7.2f}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "R3 Optimization"
        headers = ["#", "Variante", "Trades", "WR%", "PF", "Net ($)", "Rdt%", "MaxDD ($)", "AvgWin", "AvgLoss", "TSL"]
        hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        border = Border(*(Side(style='thin'),) * 4)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill, cell.font, cell.border = hf, hfont, border
        for r, s in enumerate(results_rdt, 2):
            vals = [r - 1, s['name'], s['trades'], round(s['wr'], 1), round(s['pf'], 2),
                    round(s['net'], 2), round(s['rendement'], 1), round(s['max_dd'], 2),
                    round(s.get('avg_win', 0), 2), round(s.get('avg_loss', 0), 2),
                    s.get('tsl_count', 0)]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = border
            if r == 2:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).font = Font(bold=True, color="00AA00")
        ws.column_dimensions['B'].width = 52
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(_root, f"optimizer_r3_{ts_str}.xlsx")
        wb.save(fname)
        print(f"\nRapport: {fname}")
    except Exception as e:
        print(f"\n(Excel: {e})")

    mt5.shutdown()
    print("\nTermine.")


if __name__ == "__main__":
    main()
