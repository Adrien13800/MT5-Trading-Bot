#!/usr/bin/env python3
"""
backtest_optimizer.py - Test rapide de variantes de strategie

Charge les donnees une seule fois, puis rejoue la strategie avec
differents parametres pour comparer les resultats.

Usage:
  python backtest_optimizer.py
"""

import sys
import os

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import List, Dict, Optional, Callable
from copy import deepcopy

import pandas as pd
import numpy as np

_root = os.path.dirname(os.path.abspath(__file__))
_backtest_dir = os.path.join(_root, 'backtest')
for p in [_root, _backtest_dir]:
    if p not in sys.path:
        sys.path.insert(0, p)

import MetaTrader5 as mt5
import strategy_core


# ============================================================================
# CONFIG
# ============================================================================

def load_config():
    try:
        import config
        return {
            'symbols': getattr(config, 'SYMBOLS', ['US30.cash', 'US100.cash', 'US500.cash']),
            'use_daily_preferred_symbol': getattr(config, 'USE_DAILY_PREFERRED_SYMBOL', True),
            'one_symbol_at_a_time': getattr(config, 'ONE_SYMBOL_AT_A_TIME', True),
            'preferred_symbol_by_day': getattr(config, 'PREFERRED_SYMBOL_BY_DAY', None),
            'risk_percent': getattr(config, 'RISK_PERCENT', 1.0),
            'initial_balance': getattr(config, 'INITIAL_BALANCE', 10000.0),
            'login': config.MT5_LOGIN,
            'password': config.MT5_PASSWORD,
            'server': config.MT5_SERVER,
        }
    except Exception as e:
        print(f"ERREUR config: {e}")
        sys.exit(1)


# ============================================================================
# TRADE DATACLASS
# ============================================================================

@dataclass
class SimTrade:
    symbol: str
    trade_type: strategy_core.TradeType
    entry_time: datetime
    entry_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    exit_time: Optional[datetime] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    profit: float = 0.0


# ============================================================================
# CHARGEMENT DONNEES (une seule fois)
# ============================================================================

def connect_mt5(cfg):
    if not mt5.initialize():
        for p in [r"C:\Program Files\MetaTrader\terminal64.exe",
                   r"C:\Program Files\MetaTrader 5\terminal64.exe",
                   r"C:\Program Files\MetaTrader 5 Risk Manager\terminal64.exe"]:
            if os.path.exists(p) and mt5.initialize(path=p):
                break
        else:
            print("Impossible d'initialiser MT5.")
            sys.exit(1)
    if not mt5.login(login=cfg['login'], password=cfg['password'], server=cfg['server']):
        print(f"Login echoue: {mt5.last_error()}")
        sys.exit(1)
    info = mt5.account_info()
    print(f"Connecte: {info.login} @ {info.server}")


def load_data(symbols):
    m5_data = {}
    h1_data = {}
    for sym in symbols:
        si = mt5.symbol_info(sym)
        if si is None:
            continue
        if not si.visible:
            mt5.symbol_select(sym, True)
        # M5
        best = None
        for n in [50000, 100000, 200000, 500000, 1000000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_M5, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            strategy_core.compute_indicators(df)
            m5_data[sym] = df
            print(f"  {sym}: {len(df)} barres M5")
        # H1
        best = None
        for n in [10000, 50000, 100000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_H1, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            h1_data[sym] = df
            print(f"  {sym}: {len(df)} barres H1")
    return m5_data, h1_data


def calc_lot(symbol, entry, sl, balance, risk_pct):
    si = mt5.symbol_info(symbol)
    if not si:
        return 0.0
    sd = abs(entry - sl)
    if sd <= 0:
        return 0.0
    risk = balance * (risk_pct / 100.0)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        rpl = (sd / ts) * tv
    elif si.trade_contract_size > 0:
        rpl = (sd * si.trade_contract_size) / entry
    else:
        rpl = sd * si.point
    if rpl <= 0:
        return 0.0
    lot = risk / rpl
    lot = max(si.volume_min, min(lot, si.volume_max))
    if si.volume_step > 0:
        lot = (lot // si.volume_step) * si.volume_step
    return round(lot, 2)


def calc_profit(symbol, entry, exit_p, lot, ttype):
    si = mt5.symbol_info(symbol)
    if not si:
        return 0.0
    diff = (exit_p - entry) if ttype == strategy_core.TradeType.LONG else (entry - exit_p)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        return (diff / ts) * tv * lot
    if si.trade_contract_size > 0:
        return (diff * si.trade_contract_size * lot) / entry
    return diff * si.point * lot


# ============================================================================
# VARIANTES DU FILTRE H1
# ============================================================================

def h1_filter_3bars_strict(df_h1, trade_type):
    """Original: 3 barres, rises>=2 ou falls>=2"""
    if df_h1 is None or len(df_h1) < 3:
        return False
    prices = df_h1.iloc[-3:]['close'].values
    if trade_type == strategy_core.TradeType.LONG:
        if prices[-1] < prices[0]:
            return False
        return sum(1 for i in range(1, len(prices)) if prices[i] > prices[i-1]) >= 2
    else:
        if prices[-1] > prices[0]:
            return False
        return sum(1 for i in range(1, len(prices)) if prices[i] < prices[i-1]) >= 2


def h1_filter_2bars(df_h1, trade_type):
    """2 barres: juste la derniere fermee vs l'avant-derniere"""
    if df_h1 is None or len(df_h1) < 2:
        return False
    prices = df_h1.iloc[-2:]['close'].values
    if trade_type == strategy_core.TradeType.LONG:
        return prices[1] > prices[0]
    else:
        return prices[1] < prices[0]


def h1_filter_direction_3bars(df_h1, trade_type):
    """Direction globale: close[-1] vs close[-3] (pas besoin que chaque barre aille dans le sens)"""
    if df_h1 is None or len(df_h1) < 3:
        return False
    prices = df_h1.iloc[-3:]['close'].values
    if trade_type == strategy_core.TradeType.LONG:
        return prices[-1] > prices[0]
    else:
        return prices[-1] < prices[0]


def h1_filter_direction_2bars_plus_momentum(df_h1, trade_type):
    """2 barres direction + la derniere barre doit aller dans le bon sens"""
    if df_h1 is None or len(df_h1) < 3:
        return False
    prices = df_h1.iloc[-3:]['close'].values
    if trade_type == strategy_core.TradeType.LONG:
        return prices[-1] > prices[0] and prices[-1] > prices[-2]
    else:
        return prices[-1] < prices[0] and prices[-1] < prices[-2]


def h1_filter_disabled(df_h1, trade_type):
    """Pas de filtre H1"""
    return True


# ============================================================================
# MOTEUR DE SIMULATION PARAMETRABLE
# ============================================================================

def run_variant(m5_data, h1_data, cfg,
                h1_filter_fn=h1_filter_3bars_strict,
                rr_flat=1.0, rr_trending=1.5,
                atr_sl_mult=1.5,
                cooldown_bars=0,
                use_h1=True):
    """Lance une simulation avec les parametres donnes."""

    symbols = cfg['symbols']
    use_daily_pref = cfg.get('use_daily_preferred_symbol', True)
    one_at_a_time = cfg.get('one_symbol_at_a_time', True)
    pref_by_day = cfg.get('preferred_symbol_by_day') or {}
    risk_pct = cfg.get('risk_percent', 1.0)
    balance = cfg.get('initial_balance', 10000.0)

    # Sauvegarder les constantes originales
    orig_rr_flat = strategy_core.RISK_REWARD_RATIO_FLAT
    orig_rr_trend = strategy_core.RISK_REWARD_RATIO_TRENDING
    orig_atr_sl = strategy_core.ATR_SL_MULTIPLIER
    orig_h1 = strategy_core.USE_H1_TREND_FILTER

    # Appliquer les parametres de la variante
    strategy_core.RISK_REWARD_RATIO_FLAT = rr_flat
    strategy_core.RISK_REWARD_RATIO_TRENDING = rr_trending
    strategy_core.ATR_SL_MULTIPLIER = atr_sl_mult

    # Timeline
    events = []
    for sym in symbols:
        if sym not in m5_data:
            continue
        d = m5_data[sym]
        for i in range(50, len(d)):
            ts = d.index[i]
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            events.append((ts, sym, i))
    events.sort(key=lambda x: x[0])

    open_trades: Dict[str, List[SimTrade]] = {}
    closed_trades: List[SimTrade] = []
    last_bar_time: Dict[str, datetime] = {}
    last_loss_time: Dict[str, datetime] = {}

    for current_bar_time, symbol, bar_index in events:
        df = m5_data[symbol]
        market_data = df.iloc[:bar_index + 1]
        current_bar = df.iloc[bar_index]

        # Gestion SL/TP
        if symbol in open_trades and open_trades[symbol]:
            to_close = []
            for idx, trade in enumerate(open_trades[symbol]):
                hit = None
                if trade.trade_type == strategy_core.TradeType.LONG:
                    if current_bar['low'] <= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['high'] >= trade.take_profit:
                        hit = ('TP', trade.take_profit)
                else:
                    if current_bar['high'] >= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['low'] <= trade.take_profit:
                        hit = ('TP', trade.take_profit)
                if hit:
                    trade.exit_reason, trade.exit_price = hit
                    trade.exit_time = df.index[bar_index]
                    trade.profit = calc_profit(symbol, trade.entry_price, trade.exit_price,
                                               trade.lot_size, trade.trade_type)
                    balance += trade.profit
                    if trade.profit < 0:
                        last_loss_time[symbol] = current_bar_time
                    closed_trades.append(trade)
                    to_close.append(idx)
            for idx in reversed(to_close):
                open_trades[symbol].pop(idx)
            if symbol in open_trades and not open_trades[symbol]:
                del open_trades[symbol]

        # Skip barre deja vue
        if symbol in last_bar_time and current_bar_time <= last_bar_time[symbol]:
            continue
        last_bar_time[symbol] = current_bar_time

        # Daily preferred
        if use_daily_pref and pref_by_day:
            wd = current_bar_time.weekday() if hasattr(current_bar_time, 'weekday') \
                else current_bar_time.to_pydatetime().weekday()
            pref = pref_by_day.get(wd)
            if pref is not None and symbol != pref:
                continue

        # Cooldown
        if cooldown_bars > 0 and symbol in last_loss_time:
            elapsed = (current_bar_time - last_loss_time[symbol]).total_seconds() / 300
            if elapsed < cooldown_bars:
                continue

        def has_other_open():
            if not one_at_a_time:
                return False
            for s, tl in open_trades.items():
                if s != symbol and tl:
                    return True
            return False

        # H1 data
        df_h1_filtered = strategy_core.get_h1_data_at_time(
            h1_data.get(symbol, pd.DataFrame()), current_bar_time
        ) if symbol in h1_data else None

        # Signaux
        for allow, check_fn, ttype, sl_fn, tp_sign in [
            (strategy_core.ALLOW_LONG, strategy_core.check_long_signal,
             strategy_core.TradeType.LONG, strategy_core.calculate_sl_long, 1),
            (strategy_core.ALLOW_SHORT, strategy_core.check_short_signal,
             strategy_core.TradeType.SHORT, strategy_core.calculate_sl_short, -1),
        ]:
            if not allow:
                continue

            # Verifier le signal M5 (crossover) - temporairement desactiver H1 dans strategy_core
            # car on gere le H1 nous-memes avec la variante
            old_h1_flag = strategy_core.USE_H1_TREND_FILTER
            if use_h1:
                # Desactiver dans strategy_core, on verifie manuellement
                strategy_core.USE_H1_TREND_FILTER = False

            signal = check_fn(market_data, df_h1_filtered, symbol)

            strategy_core.USE_H1_TREND_FILTER = old_h1_flag

            if not signal:
                continue

            # Appliquer notre filtre H1 custom
            if use_h1:
                if not h1_filter_fn(df_h1_filtered, ttype):
                    continue

            # Calcul entry/SL/TP
            if bar_index + 1 < len(df):
                entry_price = float(df.iloc[bar_index + 1]['open'])
            else:
                entry_price = float(current_bar['close'])

            stop_loss = sl_fn(market_data, 10)

            if ttype == strategy_core.TradeType.LONG:
                stop_dist = entry_price - stop_loss
                if stop_loss <= 0 or stop_loss >= entry_price:
                    continue
            else:
                stop_dist = stop_loss - entry_price
                if stop_loss <= 0 or stop_loss <= entry_price:
                    continue

            sl_pct = abs(entry_price - stop_loss) / entry_price if entry_price > 0 else 0
            if sl_pct > 0.05:
                continue

            rr = strategy_core.get_risk_reward_ratio(market_data)
            take_profit = entry_price + (stop_dist * rr * tp_sign)
            lot = calc_lot(symbol, entry_price, stop_loss, balance, risk_pct)

            if lot <= 0 or has_other_open():
                continue

            trade = SimTrade(
                symbol=symbol, trade_type=ttype,
                entry_time=current_bar_time,
                entry_price=entry_price, stop_loss=stop_loss,
                take_profit=take_profit, lot_size=lot,
            )
            open_trades.setdefault(symbol, []).append(trade)

    # Restaurer constantes
    strategy_core.RISK_REWARD_RATIO_FLAT = orig_rr_flat
    strategy_core.RISK_REWARD_RATIO_TRENDING = orig_rr_trend
    strategy_core.ATR_SL_MULTIPLIER = orig_atr_sl
    strategy_core.USE_H1_TREND_FILTER = orig_h1

    return closed_trades, balance


# ============================================================================
# STATS
# ============================================================================

def compute_stats(trades, initial_balance, final_balance):
    closed = [t for t in trades if t.exit_reason in ('SL', 'TP')]
    total = len(closed)
    if total == 0:
        return {'trades': 0, 'wr': 0, 'net': 0, 'pf': 0, 'rendement': 0,
                'avg_win': 0, 'avg_loss': 0, 'max_dd': 0}
    wins = [t for t in closed if t.profit > 0]
    losses = [t for t in closed if t.profit < 0]
    total_win = sum(t.profit for t in wins)
    total_loss = abs(sum(t.profit for t in losses))
    wr = len(wins) / total * 100
    pf = total_win / total_loss if total_loss > 0 else 999
    net = final_balance - initial_balance
    rendement = net / initial_balance * 100
    avg_win = total_win / len(wins) if wins else 0
    avg_loss = total_loss / len(losses) if losses else 0

    # Max drawdown
    equity = initial_balance
    peak = equity
    max_dd = 0
    for t in sorted(closed, key=lambda x: x.entry_time):
        equity += t.profit
        if equity > peak:
            peak = equity
        dd = peak - equity
        if dd > max_dd:
            max_dd = dd

    return {
        'trades': total, 'wins': len(wins), 'losses': len(losses),
        'wr': wr, 'net': net, 'pf': pf, 'rendement': rendement,
        'avg_win': avg_win, 'avg_loss': avg_loss, 'max_dd': max_dd,
    }


# ============================================================================
# MAIN
# ============================================================================

def main():
    cfg = load_config()

    print("="*70)
    print("OPTIMISEUR DE BACKTEST")
    print("="*70)

    connect_mt5(cfg)

    print("\nChargement des donnees...")
    m5_data, h1_data = load_data(cfg['symbols'])
    if not m5_data:
        print("Aucune donnee.")
        sys.exit(1)

    initial = cfg.get('initial_balance', 10000.0)

    # Definition des variantes a tester
    variants = [
        # (nom, h1_filter_fn, rr_flat, rr_trending, atr_sl_mult, cooldown, use_h1)
        ("BASELINE (H1 3bars strict)",  h1_filter_3bars_strict,               1.0, 1.5, 1.5, 0, True),
        ("H1 2 barres",                 h1_filter_2bars,                      1.0, 1.5, 1.5, 0, True),
        ("H1 direction globale",        h1_filter_direction_3bars,            1.0, 1.5, 1.5, 0, True),
        ("H1 direction + momentum",     h1_filter_direction_2bars_plus_momentum, 1.0, 1.5, 1.5, 0, True),
        ("H1 desactive",                h1_filter_disabled,                   1.0, 1.5, 1.5, 0, True),

        ("BASELINE + RR 1:2 trending",  h1_filter_3bars_strict,              1.0, 2.0, 1.5, 0, True),
        ("H1 2bars + RR 1:2 trending",  h1_filter_2bars,                     1.0, 2.0, 1.5, 0, True),
        ("H1 direction + RR 1:2",       h1_filter_direction_3bars,           1.0, 2.0, 1.5, 0, True),

        ("BASELINE + ATR SL 1.0",       h1_filter_3bars_strict,              1.0, 1.5, 1.0, 0, True),
        ("BASELINE + ATR SL 2.0",       h1_filter_3bars_strict,              1.0, 1.5, 2.0, 0, True),
        ("H1 2bars + ATR SL 2.0",       h1_filter_2bars,                     1.0, 1.5, 2.0, 0, True),

        ("BASELINE + cooldown 3",       h1_filter_3bars_strict,              1.0, 1.5, 1.5, 3, True),
        ("H1 2bars + cooldown 3",       h1_filter_2bars,                     1.0, 1.5, 1.5, 3, True),

        ("H1 2bars + RR 1:2 + ATR 2.0", h1_filter_2bars,                    1.0, 2.0, 2.0, 0, True),
        ("H1 dir + RR 1:2 + ATR 2.0",  h1_filter_direction_3bars,           1.0, 2.0, 2.0, 0, True),
        ("H1 dir+mom + RR 1:2 + cd3",  h1_filter_direction_2bars_plus_momentum, 1.0, 2.0, 1.5, 3, True),
    ]

    results = []
    print(f"\nTest de {len(variants)} variantes...\n")

    for i, (name, h1_fn, rr_f, rr_t, atr_sl, cd, use_h1) in enumerate(variants, 1):
        print(f"  [{i:2d}/{len(variants)}] {name}...", end=" ", flush=True)
        trades, final_bal = run_variant(
            m5_data, h1_data, cfg,
            h1_filter_fn=h1_fn, rr_flat=rr_f, rr_trending=rr_t,
            atr_sl_mult=atr_sl, cooldown_bars=cd, use_h1=use_h1
        )
        stats = compute_stats(trades, initial, final_bal)
        stats['name'] = name
        results.append(stats)
        print(f"{stats['trades']} trades | WR {stats['wr']:.1f}% | PF {stats['pf']:.2f} | "
              f"Rdt {stats['rendement']:.1f}% | DD {stats['max_dd']:.0f}")

    # Tri par rendement
    results.sort(key=lambda x: x['rendement'], reverse=True)

    print(f"\n{'='*110}")
    print(f"{'CLASSEMENT':^110}")
    print(f"{'='*110}")
    print(f"{'#':>3} {'Variante':<40} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8} {'AvgWin':>8} {'AvgLoss':>8}")
    print("-"*110)
    for i, s in enumerate(results, 1):
        marker = " <-- BEST" if i == 1 else ""
        print(f"{i:3d} {s['name']:<40} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$ "
              f"{s.get('avg_win',0):7.0f}$ {s.get('avg_loss',0):7.0f}${marker}")

    # Sauvegarder en Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparaison Variantes"
        headers = ["#", "Variante", "Trades", "Wins", "Losses", "WR%", "PF",
                   "Net ($)", "Rendement%", "Max DD ($)", "Avg Win", "Avg Loss"]
        hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        border = Border(*(Side(style='thin'),)*4)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
        for r, s in enumerate(results, 2):
            vals = [r-1, s['name'], s['trades'], s.get('wins',0), s.get('losses',0),
                    round(s['wr'], 1), round(s['pf'], 2), round(s['net'], 2),
                    round(s['rendement'], 1), round(s['max_dd'], 2),
                    round(s.get('avg_win',0), 2), round(s.get('avg_loss',0), 2)]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
            # Colorer la premiere ligne en vert
            if r == 2:
                for c in range(1, len(headers)+1):
                    ws.cell(row=r, column=c).font = Font(bold=True, color="00AA00")
        ws.column_dimensions['B'].width = 42
        for c in 'CDEFGHIJKL':
            ws.column_dimensions[c].width = 13
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(_root, f"optimizer_results_{ts}.xlsx")
        wb.save(fname)
        print(f"\nRapport: {fname}")
    except Exception as e:
        print(f"\n(Excel non genere: {e})")

    mt5.shutdown()
    print("\nTermine.")


if __name__ == "__main__":
    main()
