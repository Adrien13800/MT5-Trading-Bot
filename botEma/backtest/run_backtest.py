#!/usr/bin/env python3
"""
Script de lancement du backtest
Teste la stratégie sur 3 ans de données historiques
"""

import sys
import os
from datetime import datetime, timedelta
from collections import defaultdict, Counter

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERREUR: pandas et openpyxl sont requis pour generer le rapport Excel")
    print("   Installez-les avec: pip install pandas openpyxl")
    sys.exit(1)

# S'assurer que le dossier backtest est en premier dans le path (config.py = backtest/config.py)
_backtest_dir = os.path.dirname(os.path.abspath(__file__))
if _backtest_dir not in sys.path:
    sys.path.insert(0, _backtest_dir)
else:
    # Déplacer en première position pour priorité à backtest/config.py
    sys.path.remove(_backtest_dir)
    sys.path.insert(0, _backtest_dir)

from ema_mt5_bot_backtest import (
    MT5BacktestBot, BacktestStats, SimulatedTrade, TradeType, 
    RISK_REWARD_RATIO, USE_H1_TREND_FILTER, ALLOW_LONG, ALLOW_SHORT,
    TradingSession, MarketCondition, MarketTrend
)

def load_config():
    """Charge la configuration depuis config.py"""
    try:
        import config
        return {
            'login': config.MT5_LOGIN,
            'password': config.MT5_PASSWORD,
            'server': config.MT5_SERVER,
            'symbols': getattr(config, 'SYMBOLS', ['US30.cash', 'US100.cash', 'US500.cash']),
            'risk': getattr(config, 'RISK_PERCENT', 0.5),
            'max_daily_loss': getattr(config, 'MAX_DAILY_LOSS', -250.0),
            'initial_balance': getattr(config, 'INITIAL_BALANCE', 10000.0),
            'years_back': getattr(config, 'YEARS_BACK', 3),
            'use_all_available': getattr(config, 'USE_ALL_AVAILABLE_DATA', True),
            'use_daily_preferred_symbol': getattr(config, 'USE_DAILY_PREFERRED_SYMBOL', True),
            'one_symbol_at_a_time': getattr(config, 'ONE_SYMBOL_AT_A_TIME', True),
            'preferred_symbol_by_day': getattr(config, 'PREFERRED_SYMBOL_BY_DAY', None),
            'months_back': getattr(config, 'MONTHS_BACK', None),
            'use_daily_loss_in_backtest': getattr(config, 'USE_DAILY_LOSS_IN_BACKTEST', False),
            # Prix d'entrée backtest : True = open de la barre suivante (plus réaliste, proche du fill réel ask/bid)
            'use_next_bar_open_for_entry': getattr(config, 'USE_NEXT_BAR_OPEN_FOR_ENTRY', True),
        }
    except ImportError:
        print("ERREUR: Fichier config.py non trouve dans le dossier backtest")
        sys.exit(1)
    except AttributeError as e:
        print(f"ERREUR dans config.py: {e}")
        sys.exit(1)


def run_backtest_core(config_overrides=None):
    """
    Lance le backtest et retourne (bot, stats, symbol_stats).
    config_overrides: dict optionnel pour surcharger la config (ex: {'symbols': [...]}).
    """
    import MetaTrader5 as mt5
    cfg = {**load_config(), **(config_overrides or {})}

def calculate_stats(bot: MT5BacktestBot) -> BacktestStats:
    """Calcule les statistiques du backtest"""
    stats = BacktestStats()
    stats.initial_balance = bot.initial_balance
    stats.final_balance = bot.current_balance
    stats.total_trades = len(bot.closed_trades)
    
    if stats.total_trades == 0:
        return stats
    
    winning_trades = [t for t in bot.closed_trades if t.profit > 0]
    losing_trades = [t for t in bot.closed_trades if t.profit < 0]
    
    stats.winning_trades = len(winning_trades)
    stats.losing_trades = len(losing_trades)
    stats.win_rate = (stats.winning_trades / stats.total_trades) * 100 if stats.total_trades > 0 else 0
    
    stats.total_profit = sum(t.profit for t in winning_trades) if winning_trades else 0
    stats.total_loss = abs(sum(t.profit for t in losing_trades)) if losing_trades else 0
    stats.net_profit = stats.final_balance - stats.initial_balance
    stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else 0
    
    stats.avg_win = stats.total_profit / stats.winning_trades if stats.winning_trades > 0 else 0
    stats.avg_loss = stats.total_loss / stats.losing_trades if stats.losing_trades > 0 else 0
    
    stats.largest_win = max((t.profit for t in winning_trades), default=0)
    stats.largest_loss = min((t.profit for t in losing_trades), default=0)
    
    # Calculer le drawdown maximum
    if bot.equity_curve:
        peak = bot.equity_curve[0]
        max_dd = 0
        for equity in bot.equity_curve:
            if equity > peak:
                peak = equity
            dd = peak - equity
            if dd > max_dd:
                max_dd = dd
        
        stats.max_drawdown = max_dd
        stats.max_drawdown_pct = (max_dd / peak * 100) if peak > 0 else 0
    
    stats.return_pct = ((stats.final_balance - stats.initial_balance) / stats.initial_balance) * 100
    
    # Calculer les statistiques de positions simultanées
    # En analysant l'historique des trades ouverts à chaque moment
    # On va parcourir tous les trades fermés et reconstruire l'état à chaque moment
    
    # Créer une timeline de tous les événements (ouverture/fermeture)
    events = []
    for trade in bot.closed_trades:
        events.append((trade.entry_time, 'open', trade))
        events.append((trade.exit_time, 'close', trade))
    
    # Trier par date
    events.sort(key=lambda x: x[0])
    
    # Simuler l'état des positions au fil du temps
    open_positions_by_symbol = defaultdict(list)  # {symbol: [list of trades]}
    max_concurrent = 0
    times_multiple = 0
    times_long_short = 0
    
    for event_time, event_type, trade in events:
        if event_type == 'open':
            if trade.symbol not in open_positions_by_symbol:
                open_positions_by_symbol[trade.symbol] = []
            open_positions_by_symbol[trade.symbol].append(trade)
        elif event_type == 'close':
            if trade.symbol in open_positions_by_symbol:
                # Retirer le trade de la liste (trouver par entry_time et type)
                open_positions_by_symbol[trade.symbol] = [
                    t for t in open_positions_by_symbol[trade.symbol] 
                    if not (t.entry_time == trade.entry_time and t.type == trade.type)
                ]
                if len(open_positions_by_symbol[trade.symbol]) == 0:
                    del open_positions_by_symbol[trade.symbol]
        
        # Compter les positions totales ouvertes
        total_open = sum(len(trades) for trades in open_positions_by_symbol.values())
        if total_open > max_concurrent:
            max_concurrent = total_open
        
        # Compter les moments avec plusieurs positions
        if total_open > 1:
            times_multiple += 1
        
        # Compter les moments avec LONG et SHORT simultanés sur le même symbole
        for symbol, trades in open_positions_by_symbol.items():
            has_long = any(t.type == TradeType.LONG for t in trades)
            has_short = any(t.type == TradeType.SHORT for t in trades)
            if has_long and has_short:
                times_long_short += 1
                break  # Compter une seule fois par événement
    
    stats.max_concurrent_positions = max_concurrent
    stats.times_multiple_positions = times_multiple
    stats.times_long_short_simultaneous = times_long_short
    
    # Calculer les statistiques par R:R
    # Séparer les trades par R:R utilisé (1.0 ou 1.5)
    # Utiliser une tolérance de 0.01 pour gérer les arrondis
    trades_rr_1_0 = [t for t in bot.closed_trades if t.risk_reward_ratio > 0 and abs(t.risk_reward_ratio - 1.0) < 0.01]
    trades_rr_1_5 = [t for t in bot.closed_trades if t.risk_reward_ratio > 0 and abs(t.risk_reward_ratio - 1.5) < 0.01]
    
    # Statistiques pour R:R 1.0
    stats.rr_1_0_count = len(trades_rr_1_0)
    stats.rr_1_0_pct = (stats.rr_1_0_count / stats.total_trades * 100) if stats.total_trades > 0 else 0
    winning_rr_1_0 = [t for t in trades_rr_1_0 if t.profit > 0]
    stats.rr_1_0_win_rate = (len(winning_rr_1_0) / stats.rr_1_0_count * 100) if stats.rr_1_0_count > 0 else 0
    stats.rr_1_0_pnl = sum(t.profit for t in trades_rr_1_0)  # PnL total pour R:R 1.0
    
    # Statistiques pour R:R 1.5 (trades réels)
    stats.rr_1_5_count = len(trades_rr_1_5)
    stats.rr_1_5_pct = (stats.rr_1_5_count / stats.total_trades * 100) if stats.total_trades > 0 else 0
    winning_rr_1_5 = [t for t in trades_rr_1_5 if t.profit > 0]
    stats.rr_1_5_win_rate = (len(winning_rr_1_5) / stats.rr_1_5_count * 100) if stats.rr_1_5_count > 0 else 0
    stats.rr_1_5_pnl = sum(t.profit for t in trades_rr_1_5)  # PnL total pour R:R 1.5
    
    # Statistiques pour R:R 3.0 (simulation réelle : trades simulés en parallèle)
    trades_rr_3_0 = bot.closed_trades_3r  # Trades 3.0R simulés
    stats.rr_3_0_count = len(trades_rr_3_0)
    stats.rr_3_0_pct = (stats.rr_3_0_count / stats.total_trades * 100) if stats.total_trades > 0 else 0
    winning_rr_3_0 = [t for t in trades_rr_3_0 if t.profit > 0]
    stats.rr_3_0_win_rate = (len(winning_rr_3_0) / stats.rr_3_0_count * 100) if stats.rr_3_0_count > 0 else 0
    stats.rr_3_0_pnl = sum(t.profit for t in trades_rr_3_0)  # PnL total pour R:R 3.0 (simulation réelle)
    
    return stats

def print_stats(stats: BacktestStats):
    """Affiche les statistiques du backtest"""
    print("\n" + "=" * 70)
    print("RESULTATS DU BACKTEST")
    print("=" * 70)
    
    print(f"\nBalance:")
    print(f"   Initiale: {stats.initial_balance:.2f}")
    print(f"   Finale: {stats.final_balance:.2f}")
    print(f"   Profit net: {stats.net_profit:.2f} ({stats.return_pct:.2f}%)")
    
    print(f"\nTrades:")
    print(f"   Total: {stats.total_trades}")
    print(f"   Gagnants: {stats.winning_trades} ({stats.win_rate:.2f}%)")
    print(f"   Perdants: {stats.losing_trades} ({100 - stats.win_rate:.2f}%)")
    
    print(f"\n[PERFORMANCE]")
    print(f"   Profit total: {stats.total_profit:.2f}")
    print(f"   Perte total: {stats.total_loss:.2f}")
    print(f"   Profit Factor: {stats.profit_factor:.2f}")
    print(f"   Gain moyen: {stats.avg_win:.2f}")
    print(f"   Perte moyenne: {stats.avg_loss:.2f}")
    print(f"   Plus gros gain: {stats.largest_win:.2f}")
    print(f"   Plus grosse perte: {stats.largest_loss:.2f}")
    
    print(f"\n[RISQUE]")
    print(f"   Drawdown max: {stats.max_drawdown:.2f} ({stats.max_drawdown_pct:.2f}%)")
    
    print(f"\n[POSITIONS SIMULTANEES]")
    print(f"   Max positions simultanées: {stats.max_concurrent_positions}")
    print(f"   Moments avec plusieurs positions: {stats.times_multiple_positions}")
    print(f"   Moments LONG+SHORT simultanés: {stats.times_long_short_simultaneous}")
    
    print(f"\n[STATISTIQUES PAR R:R]")
    print(f"   R:R 1.0:")
    print(f"      Nombre de trades: {stats.rr_1_0_count} ({stats.rr_1_0_pct:.2f}%)")
    print(f"      Win Rate: {stats.rr_1_0_win_rate:.2f}%")
    print(f"      PnL: {stats.rr_1_0_pnl:.2f} USD")
    print(f"   R:R 1.5 (réel):")
    print(f"      Nombre de trades: {stats.rr_1_5_count} ({stats.rr_1_5_pct:.2f}%)")
    print(f"      Win Rate: {stats.rr_1_5_win_rate:.2f}%")
    print(f"      PnL: {stats.rr_1_5_pnl:.2f} USD")
    print(f"   R:R 3.0 (simulé précis):")
    print(f"      Nombre de trades: {stats.rr_3_0_count} ({stats.rr_3_0_pct:.2f}%)")
    print(f"      Win Rate: {stats.rr_3_0_win_rate:.2f}%")
    print(f"      PnL: {stats.rr_3_0_pnl:.2f} USD")
    
    print("\n" + "=" * 70)


def print_advanced_analytics(bot: MT5BacktestBot):
    """Affiche les analytics avancées dans la console"""
    print("\n" + "=" * 70)
    print("ANALYTICS AVANCEES")
    print("=" * 70)
    
    if not bot.closed_trades:
        print("   Aucun trade à analyser")
        return
    
    # ========== PAR SESSION ==========
    print("\n[PERFORMANCE PAR SESSION DE TRADING]")
    print("-" * 50)
    
    session_data = {}
    for session in TradingSession:
        trades = [t for t in bot.closed_trades if t.session == session]
        if trades:
            winning = len([t for t in trades if t.profit > 0])
            total = len(trades)
            wr = (winning / total * 100) if total > 0 else 0
            net = sum(t.profit for t in trades)
            session_data[session] = {'trades': total, 'wr': wr, 'net': net}
    
    if session_data:
        best_session = max(session_data.keys(), key=lambda s: session_data[s]['wr'])
        for session, data in sorted(session_data.items(), key=lambda x: x[1]['wr'], reverse=True):
            star = " ⭐" if session == best_session else ""
            print(f"   {session.value:10} | {data['trades']:3} trades | WR: {data['wr']:5.1f}% | Net: {data['net']:+8.2f} USD{star}")
    
    # ========== BULL vs BEAR ==========
    print("\n[PERFORMANCE BULL vs BEAR]")
    print("-" * 50)
    
    for condition in MarketCondition:
        trades = [t for t in bot.closed_trades if t.market_condition == condition]
        if trades:
            winning = len([t for t in trades if t.profit > 0])
            total = len(trades)
            wr = (winning / total * 100) if total > 0 else 0
            net = sum(t.profit for t in trades)
            label = "📈 BULL" if condition == MarketCondition.BULL else "📉 BEAR"
            print(f"   {label:10} | {total:3} trades | WR: {wr:5.1f}% | Net: {net:+8.2f} USD")
    
    # ========== TRENDING vs RANGING ==========
    print("\n[PERFORMANCE TRENDING vs RANGING]")
    print("-" * 50)
    
    for trend in MarketTrend:
        trades = [t for t in bot.closed_trades if t.market_trend == trend]
        if trades:
            winning = len([t for t in trades if t.profit > 0])
            total = len(trades)
            wr = (winning / total * 100) if total > 0 else 0
            net = sum(t.profit for t in trades)
            label = "📊 TREND" if trend == MarketTrend.TRENDING else "↔️ RANGE"
            print(f"   {label:10} | {total:3} trades | WR: {wr:5.1f}% | Net: {net:+8.2f} USD")
    
    # ========== PAR JOUR ==========
    print("\n[PERFORMANCE PAR JOUR DE LA SEMAINE]")
    print("-" * 50)
    
    day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    day_data = {}
    for day_num in range(7):
        trades = [t for t in bot.closed_trades if t.day_of_week == day_num]
        if trades:
            winning = len([t for t in trades if t.profit > 0])
            total = len(trades)
            wr = (winning / total * 100) if total > 0 else 0
            net = sum(t.profit for t in trades)
            day_data[day_num] = {'name': day_names[day_num], 'trades': total, 'wr': wr, 'net': net}
    
    if day_data:
        best_day = max(day_data.keys(), key=lambda d: day_data[d]['wr'])
        for day_num in sorted(day_data.keys()):
            data = day_data[day_num]
            star = " ⭐" if day_num == best_day else ""
            print(f"   {data['name']:10} | {data['trades']:3} trades | WR: {data['wr']:5.1f}% | Net: {data['net']:+8.2f} USD{star}")
    
    # ========== MEILLEURE COMBINAISON ==========
    print("\n[TOP 5 MEILLEURES COMBINAISONS (min 5 trades)]")
    print("-" * 50)
    
    combo_stats = {}
    for trade in bot.closed_trades:
        if trade.session and trade.market_condition and trade.market_trend:
            key = (trade.session, trade.market_condition, trade.market_trend)
            if key not in combo_stats:
                combo_stats[key] = {'winning': 0, 'losing': 0, 'profit': 0}
            if trade.profit > 0:
                combo_stats[key]['winning'] += 1
            else:
                combo_stats[key]['losing'] += 1
            combo_stats[key]['profit'] += trade.profit
    
    combo_results = []
    for key, s in combo_stats.items():
        total = s['winning'] + s['losing']
        if total >= 5:
            wr = (s['winning'] / total * 100) if total > 0 else 0
            combo_results.append({
                'combo': f"{key[0].value} + {key[1].value} + {key[2].value}",
                'trades': total,
                'wr': wr,
                'net': s['profit']
            })
    
    combo_results.sort(key=lambda x: x['wr'], reverse=True)
    
    for i, combo in enumerate(combo_results[:5], 1):
        print(f"   {i}. {combo['combo']}")
        print(f"      {combo['trades']} trades | WR: {combo['wr']:.1f}% | Net: {combo['net']:+.2f} USD")
    
    if not combo_results:
        print("   Pas assez de données (min 5 trades par combinaison)")
    
    # ========== RECOMMANDATIONS ==========
    print("\n[RECOMMANDATIONS]")
    print("-" * 50)
    
    recommendations = []
    
    # Session recommandée
    if session_data:
        best_sess = max(session_data.keys(), key=lambda s: session_data[s]['wr'])
        worst_sess = min(session_data.keys(), key=lambda s: session_data[s]['wr'])
        if session_data[best_sess]['wr'] - session_data[worst_sess]['wr'] > 10:
            recommendations.append(f"   ✅ Privilégier la session {best_sess.value} (WR: {session_data[best_sess]['wr']:.1f}%)")
            recommendations.append(f"   ❌ Éviter la session {worst_sess.value} (WR: {session_data[worst_sess]['wr']:.1f}%)")
    
    # Bull vs Bear
    bull_trades = [t for t in bot.closed_trades if t.market_condition == MarketCondition.BULL]
    bear_trades = [t for t in bot.closed_trades if t.market_condition == MarketCondition.BEAR]
    if bull_trades and bear_trades:
        bull_wr = len([t for t in bull_trades if t.profit > 0]) / len(bull_trades) * 100
        bear_wr = len([t for t in bear_trades if t.profit > 0]) / len(bear_trades) * 100
        if abs(bull_wr - bear_wr) > 10:
            if bull_wr > bear_wr:
                recommendations.append(f"   ✅ Meilleure performance en marché BULL ({bull_wr:.1f}% vs {bear_wr:.1f}%)")
            else:
                recommendations.append(f"   ✅ Meilleure performance en marché BEAR ({bear_wr:.1f}% vs {bull_wr:.1f}%)")
    
    # Trending vs Ranging
    trend_trades = [t for t in bot.closed_trades if t.market_trend == MarketTrend.TRENDING]
    range_trades = [t for t in bot.closed_trades if t.market_trend == MarketTrend.RANGING]
    if trend_trades and range_trades:
        trend_wr = len([t for t in trend_trades if t.profit > 0]) / len(trend_trades) * 100
        range_wr = len([t for t in range_trades if t.profit > 0]) / len(range_trades) * 100
        if abs(trend_wr - range_wr) > 10:
            if trend_wr > range_wr:
                recommendations.append(f"   ✅ Meilleure performance en marché TRENDING ({trend_wr:.1f}% vs {range_wr:.1f}%)")
            else:
                recommendations.append(f"   ✅ Meilleure performance en marché RANGING ({range_wr:.1f}% vs {trend_wr:.1f}%)")
    
    if recommendations:
        for rec in recommendations:
            print(rec)
    else:
        print("   Pas de différence significative détectée entre les conditions")
    
    print("\n" + "=" * 70)

def generate_excel_report(bot: MT5BacktestBot, stats: BacktestStats, output_file: str = "backtest_report.xlsx", period_start=None, period_end=None):
    """Génère un rapport Excel détaillé avec toutes les statistiques.
    Si period_start et period_end sont fournis, l'onglet Par Mois affiche TOUS les mois de la période (même 0 trade)."""
    print(f"\nGeneration du rapport Excel: {output_file}...")
    
    wb = Workbook()
    
    # ========== ONGLET 1: RÉSUMÉ GLOBAL ==========
    ws_summary = wb.active
    ws_summary.title = "Résumé Global"
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws_summary['A1'] = "RAPPORT DE BACKTEST - Stratégie EMA 20 / SMA 50 (Croisement)"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    row = 3
    
    # Section Balance
    ws_summary[f'A{row}'] = "BALANCE"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    ws_summary[f'A{row}'] = "Balance Initiale"
    ws_summary[f'B{row}'] = f"{stats.initial_balance:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Balance Finale"
    ws_summary[f'B{row}'] = f"{stats.final_balance:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Profit Net"
    ws_summary[f'B{row}'] = f"{stats.net_profit:.2f} USD"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.net_profit > 0 else "AA0000")
    row += 1
    
    ws_summary[f'A{row}'] = "Rendement"
    ws_summary[f'B{row}'] = f"{stats.return_pct:.2f}%"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.return_pct > 0 else "AA0000")
    row += 2
    
    # Section Trades
    ws_summary[f'A{row}'] = "TRADES"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    ws_summary[f'A{row}'] = "Total Trades"
    ws_summary[f'B{row}'] = stats.total_trades
    row += 1
    
    ws_summary[f'A{row}'] = "Trades Gagnants"
    ws_summary[f'B{row}'] = stats.winning_trades
    row += 1
    
    ws_summary[f'A{row}'] = "Trades Perdants"
    ws_summary[f'B{row}'] = stats.losing_trades
    row += 1
    
    ws_summary[f'A{row}'] = "WIN RATE (WR)"
    ws_summary[f'B{row}'] = f"{stats.win_rate:.2f}%"
    ws_summary[f'B{row}'].font = Font(bold=True, size=12, color="00AA00" if stats.win_rate >= 50 else "AA0000")
    row += 2
    
    # Section Performance
    ws_summary[f'A{row}'] = "PERFORMANCE"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    ws_summary[f'A{row}'] = "Profit Total"
    ws_summary[f'B{row}'] = f"{stats.total_profit:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Perte Total"
    ws_summary[f'B{row}'] = f"{stats.total_loss:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Profit Factor"
    ws_summary[f'B{row}'] = f"{stats.profit_factor:.2f}"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.profit_factor > 1 else "AA0000")
    row += 1
    
    ws_summary[f'A{row}'] = "Gain Moyen"
    ws_summary[f'B{row}'] = f"{stats.avg_win:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Perte Moyenne"
    ws_summary[f'B{row}'] = f"{stats.avg_loss:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Plus Gros Gain"
    ws_summary[f'B{row}'] = f"{stats.largest_win:.2f} USD"
    row += 1
    
    ws_summary[f'A{row}'] = "Plus Grosse Perte"
    ws_summary[f'B{row}'] = f"{stats.largest_loss:.2f} USD"
    row += 2
    
    # Section Risque
    ws_summary[f'A{row}'] = "RISQUE"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    ws_summary[f'A{row}'] = "Drawdown Maximum (USD)"
    ws_summary[f'B{row}'] = f"{stats.max_drawdown:.2f} USD"
    ws_summary[f'B{row}'].font = Font(bold=True, color="AA0000")
    row += 1
    
    ws_summary[f'A{row}'] = "Drawdown Maximum (%)"
    ws_summary[f'B{row}'] = f"{stats.max_drawdown_pct:.2f}%"
    ws_summary[f'B{row}'].font = Font(bold=True, color="AA0000")
    row += 2
    
    # Section Statistiques par R:R - Tableau comparatif
    ws_summary[f'A{row}'] = "COMPARAISON R:R 1.5 vs 3.0"
    ws_summary[f'A{row}'].font = title_font
    ws_summary.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # En-têtes du tableau
    ws_summary[f'A{row}'] = "Métrique"
    ws_summary[f'B{row}'] = "R:R 1.5 (Réel)"
    ws_summary[f'C{row}'] = "R:R 3.0 (Simulé)"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].font = Font(bold=True)
        ws_summary[f'{col}{row}'].fill = header_fill
        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # Nombre de trades
    ws_summary[f'A{row}'] = "Nombre de trades"
    ws_summary[f'B{row}'] = f"{stats.rr_1_5_count} ({stats.rr_1_5_pct:.2f}%)"
    ws_summary[f'C{row}'] = f"{stats.rr_3_0_count} ({stats.rr_3_0_pct:.2f}%)"
    row += 1
    
    # Win Rate
    ws_summary[f'A{row}'] = "Win Rate"
    ws_summary[f'B{row}'] = f"{stats.rr_1_5_win_rate:.2f}%"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.rr_1_5_win_rate >= 50 else "AA0000")
    ws_summary[f'C{row}'] = f"{stats.rr_3_0_win_rate:.2f}%"
    ws_summary[f'C{row}'].font = Font(bold=True, color="00AA00" if stats.rr_3_0_win_rate >= 50 else "AA0000")
    row += 1
    
    # PnL
    ws_summary[f'A{row}'] = "PnL Total"
    ws_summary[f'B{row}'] = f"{stats.rr_1_5_pnl:.2f} USD"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.rr_1_5_pnl > 0 else "AA0000")
    ws_summary[f'C{row}'] = f"{stats.rr_3_0_pnl:.2f} USD"
    ws_summary[f'C{row}'].font = Font(bold=True, color="00AA00" if stats.rr_3_0_pnl > 0 else "AA0000")
    row += 2
    
    # Section R:R 1.0 (pour référence)
    ws_summary[f'A{row}'] = "R:R 1.0 (SMA50 plate)"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    ws_summary[f'A{row}'] = "R:R 1.0 - Nombre de trades"
    ws_summary[f'B{row}'] = f"{stats.rr_1_0_count} ({stats.rr_1_0_pct:.2f}%)"
    row += 1
    
    ws_summary[f'A{row}'] = "R:R 1.0 - Win Rate"
    ws_summary[f'B{row}'] = f"{stats.rr_1_0_win_rate:.2f}%"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.rr_1_0_win_rate >= 50 else "AA0000")
    row += 1
    
    ws_summary[f'A{row}'] = "R:R 1.0 - PnL"
    ws_summary[f'B{row}'] = f"{stats.rr_1_0_pnl:.2f} USD"
    ws_summary[f'B{row}'].font = Font(bold=True, color="00AA00" if stats.rr_1_0_pnl > 0 else "AA0000")
    
    # Ajuster la largeur des colonnes
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ========== ONGLET 2: STATISTIQUES PAR MOIS ==========
    ws_monthly = wb.create_sheet("Par Mois")
    
    # En-têtes
    headers = ["Mois", "Année", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Balance Fin Mois"]
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Calculer les stats par mois
    monthly_stats = defaultdict(lambda: {'trades': []})
    for trade in bot.closed_trades:
        # Gérer Timestamp pandas ou datetime
        t = trade.exit_time
        if hasattr(t, 'to_pydatetime'):
            t = t.to_pydatetime()
        year, month = t.year, t.month
        monthly_stats[(year, month)]['trades'].append(trade)
    
    # Couvrir TOUTE la période du backtest (tous les mois, même sans trade)
    if period_start is not None and period_end is not None:
        start_ts = pd.Timestamp(period_start)
        end_ts = pd.Timestamp(period_end)
        all_months = []
        y, m = start_ts.year, start_ts.month
        end_y, end_m = end_ts.year, end_ts.month
        while (y, m) <= (end_y, end_m):
            all_months.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1
        sorted_months = all_months
    else:
        sorted_months = sorted(monthly_stats.keys())
    
    row = 2
    current_balance = bot.initial_balance
    for year, month in sorted_months:
        trades = monthly_stats.get((year, month), {}).get('trades', [])
        winning = [t for t in trades if t.profit > 0]
        losing = [t for t in trades if t.profit < 0]
        
        total_profit = sum(t.profit for t in winning)
        total_loss = abs(sum(t.profit for t in losing))
        net = total_profit - total_loss
        wr = (len(winning) / len(trades) * 100) if trades else 0
        
        # Mettre à jour la balance
        current_balance += net
        
        ws_monthly.cell(row=row, column=1).value = f"{month:02d}"
        ws_monthly.cell(row=row, column=2).value = year
        ws_monthly.cell(row=row, column=3).value = len(trades)
        ws_monthly.cell(row=row, column=4).value = len(winning)
        ws_monthly.cell(row=row, column=5).value = len(losing)
        ws_monthly.cell(row=row, column=6).value = f"{wr:.2f}%"
        ws_monthly.cell(row=row, column=7).value = f"{total_profit:.2f}"
        ws_monthly.cell(row=row, column=8).value = f"{total_loss:.2f}"
        ws_monthly.cell(row=row, column=9).value = f"{net:.2f}"
        ws_monthly.cell(row=row, column=9).font = Font(color="00AA00" if net > 0 else "AA0000")
        ws_monthly.cell(row=row, column=10).value = f"{current_balance:.2f}"
        
        # Appliquer les bordures
        for col in range(1, len(headers) + 1):
            ws_monthly.cell(row=row, column=col).border = border
        
        row += 1
    
    # Ajuster les colonnes
    for col in range(1, len(headers) + 1):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 15
    
    # ========== ONGLET 3: STATISTIQUES PAR ANNÉE ==========
    ws_yearly = wb.create_sheet("Par Année")
    
    headers = ["Année", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Balance Fin Année", "Rendement (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws_yearly.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Calculer les stats par année
    yearly_stats = defaultdict(lambda: {'trades': []})
    
    for trade in bot.closed_trades:
        year = trade.exit_time.year
        yearly_stats[year]['trades'].append(trade)
    
    sorted_years = sorted(yearly_stats.keys())
    
    row = 2
    balance_start_year = bot.initial_balance
    
    for year in sorted_years:
        trades = yearly_stats[year]['trades']
        winning = [t for t in trades if t.profit > 0]
        losing = [t for t in trades if t.profit < 0]
        
        total_profit = sum(t.profit for t in winning)
        total_loss = abs(sum(t.profit for t in losing))
        net = total_profit - total_loss
        wr = (len(winning) / len(trades) * 100) if trades else 0
        
        balance_end_year = balance_start_year + net
        return_pct = ((balance_end_year - balance_start_year) / balance_start_year * 100) if balance_start_year > 0 else 0
        
        ws_yearly.cell(row=row, column=1).value = year
        ws_yearly.cell(row=row, column=2).value = len(trades)
        ws_yearly.cell(row=row, column=3).value = len(winning)
        ws_yearly.cell(row=row, column=4).value = len(losing)
        ws_yearly.cell(row=row, column=5).value = f"{wr:.2f}%"
        ws_yearly.cell(row=row, column=6).value = f"{total_profit:.2f}"
        ws_yearly.cell(row=row, column=7).value = f"{total_loss:.2f}"
        ws_yearly.cell(row=row, column=8).value = f"{net:.2f}"
        ws_yearly.cell(row=row, column=8).font = Font(bold=True, color="00AA00" if net > 0 else "AA0000")
        ws_yearly.cell(row=row, column=9).value = f"{balance_end_year:.2f}"
        ws_yearly.cell(row=row, column=10).value = f"{return_pct:.2f}%"
        ws_yearly.cell(row=row, column=10).font = Font(bold=True, color="00AA00" if return_pct > 0 else "AA0000")
        
        for col in range(1, len(headers) + 1):
            ws_yearly.cell(row=row, column=col).border = border
        
        balance_start_year = balance_end_year
        row += 1
    
    for col in range(1, len(headers) + 1):
        ws_yearly.column_dimensions[get_column_letter(col)].width = 15
    
    # ========== ONGLET 4: DÉTAIL DES TRADES ==========
    ws_trades = wb.create_sheet("Détail Trades")
    
    headers = ["#", "Date Entrée", "Date Sortie", "Symbole", "Type", "Entry", "SL", "TP", "Exit", "Lot", "Profit (USD)", "Durée", "Raison", "Session", "Bull/Bear", "Trend/Range", "Jour"]
    for col, header in enumerate(headers, 1):
        cell = ws_trades.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    day_names = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    
    row = 2
    for idx, trade in enumerate(bot.closed_trades, 1):
        duration = (trade.exit_time - trade.entry_time).total_seconds() / 3600 if trade.exit_time else 0  # en heures
        
        ws_trades.cell(row=row, column=1).value = idx
        ws_trades.cell(row=row, column=2).value = trade.entry_time.strftime("%Y-%m-%d %H:%M")
        ws_trades.cell(row=row, column=3).value = trade.exit_time.strftime("%Y-%m-%d %H:%M") if trade.exit_time else ""
        ws_trades.cell(row=row, column=4).value = trade.symbol
        ws_trades.cell(row=row, column=5).value = trade.type.value
        ws_trades.cell(row=row, column=6).value = f"{trade.entry_price:.2f}"
        ws_trades.cell(row=row, column=7).value = f"{trade.stop_loss:.2f}"
        ws_trades.cell(row=row, column=8).value = f"{trade.take_profit:.2f}"
        ws_trades.cell(row=row, column=9).value = f"{trade.exit_price:.2f}" if trade.exit_price else ""
        ws_trades.cell(row=row, column=10).value = f"{trade.lot_size:.2f}"
        ws_trades.cell(row=row, column=11).value = f"{trade.profit:.2f}"
        ws_trades.cell(row=row, column=11).font = Font(bold=True, color="00AA00" if trade.profit > 0 else "AA0000")
        ws_trades.cell(row=row, column=12).value = f"{duration:.1f}h"
        ws_trades.cell(row=row, column=13).value = trade.exit_reason
        # Analytics avancées
        ws_trades.cell(row=row, column=14).value = trade.session.value if trade.session else ""
        ws_trades.cell(row=row, column=15).value = trade.market_condition.value if trade.market_condition else ""
        ws_trades.cell(row=row, column=16).value = trade.market_trend.value if trade.market_trend else ""
        ws_trades.cell(row=row, column=17).value = day_names[trade.day_of_week] if trade.day_of_week is not None else ""
        
        for col in range(1, len(headers) + 1):
            ws_trades.cell(row=row, column=col).border = border
        
        row += 1
    
    for col in range(1, len(headers) + 1):
        ws_trades.column_dimensions[get_column_letter(col)].width = 15
    
    # ========== ONGLET 5: ANALYTICS PAR SESSION ==========
    ws_sessions = wb.create_sheet("Par Session")
    
    headers = ["Session", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Profit Factor", "Meilleure Session?"]
    for col, header in enumerate(headers, 1):
        cell = ws_sessions.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Calculer les stats par session
    session_stats = {}
    for session in TradingSession:
        session_trades = [t for t in bot.closed_trades if t.session == session]
        if session_trades:
            winning = [t for t in session_trades if t.profit > 0]
            losing = [t for t in session_trades if t.profit < 0]
            total_profit = sum(t.profit for t in winning)
            total_loss = abs(sum(t.profit for t in losing))
            wr = (len(winning) / len(session_trades) * 100) if session_trades else 0
            pf = total_profit / total_loss if total_loss > 0 else 0
            net = total_profit - total_loss
            session_stats[session] = {
                'trades': len(session_trades),
                'winning': len(winning),
                'losing': len(losing),
                'wr': wr,
                'profit': total_profit,
                'loss': total_loss,
                'net': net,
                'pf': pf
            }
    
    # Trouver la meilleure session (par WR)
    best_session = max(session_stats.keys(), key=lambda s: session_stats[s]['wr']) if session_stats else None
    
    row = 2
    for session in TradingSession:
        if session in session_stats:
            s = session_stats[session]
            is_best = session == best_session
            ws_sessions.cell(row=row, column=1).value = session.value
            ws_sessions.cell(row=row, column=2).value = s['trades']
            ws_sessions.cell(row=row, column=3).value = s['winning']
            ws_sessions.cell(row=row, column=4).value = s['losing']
            ws_sessions.cell(row=row, column=5).value = f"{s['wr']:.2f}%"
            ws_sessions.cell(row=row, column=5).font = Font(bold=True, color="00AA00" if s['wr'] >= 50 else "AA0000")
            ws_sessions.cell(row=row, column=6).value = f"{s['profit']:.2f}"
            ws_sessions.cell(row=row, column=7).value = f"{s['loss']:.2f}"
            ws_sessions.cell(row=row, column=8).value = f"{s['net']:.2f}"
            ws_sessions.cell(row=row, column=8).font = Font(color="00AA00" if s['net'] > 0 else "AA0000")
            ws_sessions.cell(row=row, column=9).value = f"{s['pf']:.2f}"
            ws_sessions.cell(row=row, column=10).value = "⭐ OUI" if is_best else ""
            ws_sessions.cell(row=row, column=10).font = Font(bold=True, color="FFD700") if is_best else Font()
            
            for col in range(1, len(headers) + 1):
                ws_sessions.cell(row=row, column=col).border = border
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_sessions.column_dimensions[get_column_letter(col)].width = 15
    
    # ========== ONGLET 6: ANALYTICS BULL vs BEAR ==========
    ws_market = wb.create_sheet("Bull vs Bear")
    
    headers = ["Condition", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Profit Factor", "Meilleure?"]
    for col, header in enumerate(headers, 1):
        cell = ws_market.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Stats par condition de marché
    market_stats = {}
    for condition in MarketCondition:
        cond_trades = [t for t in bot.closed_trades if t.market_condition == condition]
        if cond_trades:
            winning = [t for t in cond_trades if t.profit > 0]
            losing = [t for t in cond_trades if t.profit < 0]
            total_profit = sum(t.profit for t in winning)
            total_loss = abs(sum(t.profit for t in losing))
            wr = (len(winning) / len(cond_trades) * 100) if cond_trades else 0
            pf = total_profit / total_loss if total_loss > 0 else 0
            net = total_profit - total_loss
            market_stats[condition] = {
                'trades': len(cond_trades),
                'winning': len(winning),
                'losing': len(losing),
                'wr': wr,
                'profit': total_profit,
                'loss': total_loss,
                'net': net,
                'pf': pf
            }
    
    best_market = max(market_stats.keys(), key=lambda c: market_stats[c]['wr']) if market_stats else None
    
    row = 2
    for condition in MarketCondition:
        if condition in market_stats:
            s = market_stats[condition]
            is_best = condition == best_market
            label = "📈 BULL (prix > SMA50)" if condition == MarketCondition.BULL else "📉 BEAR (prix < SMA50)"
            ws_market.cell(row=row, column=1).value = label
            ws_market.cell(row=row, column=2).value = s['trades']
            ws_market.cell(row=row, column=3).value = s['winning']
            ws_market.cell(row=row, column=4).value = s['losing']
            ws_market.cell(row=row, column=5).value = f"{s['wr']:.2f}%"
            ws_market.cell(row=row, column=5).font = Font(bold=True, color="00AA00" if s['wr'] >= 50 else "AA0000")
            ws_market.cell(row=row, column=6).value = f"{s['profit']:.2f}"
            ws_market.cell(row=row, column=7).value = f"{s['loss']:.2f}"
            ws_market.cell(row=row, column=8).value = f"{s['net']:.2f}"
            ws_market.cell(row=row, column=8).font = Font(color="00AA00" if s['net'] > 0 else "AA0000")
            ws_market.cell(row=row, column=9).value = f"{s['pf']:.2f}"
            ws_market.cell(row=row, column=10).value = "⭐ OUI" if is_best else ""
            
            for col in range(1, len(headers) + 1):
                ws_market.cell(row=row, column=col).border = border
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_market.column_dimensions[get_column_letter(col)].width = 22
    
    # ========== ONGLET 7: ANALYTICS TRENDING vs RANGING ==========
    ws_trend = wb.create_sheet("Trending vs Ranging")
    
    headers = ["Tendance", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Profit Factor", "Meilleure?"]
    for col, header in enumerate(headers, 1):
        cell = ws_trend.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Stats par tendance de marché
    trend_stats = {}
    for trend in MarketTrend:
        trend_trades = [t for t in bot.closed_trades if t.market_trend == trend]
        if trend_trades:
            winning = [t for t in trend_trades if t.profit > 0]
            losing = [t for t in trend_trades if t.profit < 0]
            total_profit = sum(t.profit for t in winning)
            total_loss = abs(sum(t.profit for t in losing))
            wr = (len(winning) / len(trend_trades) * 100) if trend_trades else 0
            pf = total_profit / total_loss if total_loss > 0 else 0
            net = total_profit - total_loss
            trend_stats[trend] = {
                'trades': len(trend_trades),
                'winning': len(winning),
                'losing': len(losing),
                'wr': wr,
                'profit': total_profit,
                'loss': total_loss,
                'net': net,
                'pf': pf
            }
    
    best_trend = max(trend_stats.keys(), key=lambda t: trend_stats[t]['wr']) if trend_stats else None
    
    row = 2
    for trend in MarketTrend:
        if trend in trend_stats:
            s = trend_stats[trend]
            is_best = trend == best_trend
            label = "📊 TRENDING (SMA50 penche)" if trend == MarketTrend.TRENDING else "↔️ RANGING (SMA50 plate)"
            ws_trend.cell(row=row, column=1).value = label
            ws_trend.cell(row=row, column=2).value = s['trades']
            ws_trend.cell(row=row, column=3).value = s['winning']
            ws_trend.cell(row=row, column=4).value = s['losing']
            ws_trend.cell(row=row, column=5).value = f"{s['wr']:.2f}%"
            ws_trend.cell(row=row, column=5).font = Font(bold=True, color="00AA00" if s['wr'] >= 50 else "AA0000")
            ws_trend.cell(row=row, column=6).value = f"{s['profit']:.2f}"
            ws_trend.cell(row=row, column=7).value = f"{s['loss']:.2f}"
            ws_trend.cell(row=row, column=8).value = f"{s['net']:.2f}"
            ws_trend.cell(row=row, column=8).font = Font(color="00AA00" if s['net'] > 0 else "AA0000")
            ws_trend.cell(row=row, column=9).value = f"{s['pf']:.2f}"
            ws_trend.cell(row=row, column=10).value = "⭐ OUI" if is_best else ""
            
            for col in range(1, len(headers) + 1):
                ws_trend.cell(row=row, column=col).border = border
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_trend.column_dimensions[get_column_letter(col)].width = 25
    
    # ========== ONGLET 8: ANALYTICS PAR JOUR DE LA SEMAINE ==========
    ws_days = wb.create_sheet("Par Jour")
    
    headers = ["Jour", "Trades", "Gagnants", "Perdants", "WR (%)", "Profit (USD)", "Perte (USD)", "Net (USD)", "Profit Factor"]
    for col, header in enumerate(headers, 1):
        cell = ws_days.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    
    row = 2
    for day_num in range(7):
        day_trades = [t for t in bot.closed_trades if t.day_of_week == day_num]
        if day_trades:
            winning = [t for t in day_trades if t.profit > 0]
            losing = [t for t in day_trades if t.profit < 0]
            total_profit = sum(t.profit for t in winning)
            total_loss = abs(sum(t.profit for t in losing))
            wr = (len(winning) / len(day_trades) * 100) if day_trades else 0
            pf = total_profit / total_loss if total_loss > 0 else 0
            net = total_profit - total_loss
            
            ws_days.cell(row=row, column=1).value = day_names[day_num]
            ws_days.cell(row=row, column=2).value = len(day_trades)
            ws_days.cell(row=row, column=3).value = len(winning)
            ws_days.cell(row=row, column=4).value = len(losing)
            ws_days.cell(row=row, column=5).value = f"{wr:.2f}%"
            ws_days.cell(row=row, column=5).font = Font(bold=True, color="00AA00" if wr >= 50 else "AA0000")
            ws_days.cell(row=row, column=6).value = f"{total_profit:.2f}"
            ws_days.cell(row=row, column=7).value = f"{total_loss:.2f}"
            ws_days.cell(row=row, column=8).value = f"{net:.2f}"
            ws_days.cell(row=row, column=8).font = Font(color="00AA00" if net > 0 else "AA0000")
            ws_days.cell(row=row, column=9).value = f"{pf:.2f}"
            
            for col in range(1, len(headers) + 1):
                ws_days.cell(row=row, column=col).border = border
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_days.column_dimensions[get_column_letter(col)].width = 15
    
    # ========== ONGLET 9: COMBINAISONS GAGNANTES ==========
    ws_combos = wb.create_sheet("Meilleures Combinaisons")
    
    # En-tête explicatif
    ws_combos['A1'] = "ANALYSE DES MEILLEURES COMBINAISONS"
    ws_combos['A1'].font = Font(bold=True, size=14)
    ws_combos.merge_cells('A1:F1')
    
    ws_combos['A3'] = "Cette analyse identifie les conditions optimales pour trader"
    ws_combos['A3'].font = Font(italic=True)
    
    # Analyser toutes les combinaisons Session + MarketCondition + MarketTrend
    combo_stats = {}
    for trade in bot.closed_trades:
        if trade.session and trade.market_condition and trade.market_trend:
            key = (trade.session, trade.market_condition, trade.market_trend)
            if key not in combo_stats:
                combo_stats[key] = {'winning': 0, 'losing': 0, 'profit': 0, 'loss': 0}
            if trade.profit > 0:
                combo_stats[key]['winning'] += 1
                combo_stats[key]['profit'] += trade.profit
            else:
                combo_stats[key]['losing'] += 1
                combo_stats[key]['loss'] += abs(trade.profit)
    
    # Calculer WR et trier
    combo_results = []
    for key, s in combo_stats.items():
        total = s['winning'] + s['losing']
        if total >= 5:  # Minimum 5 trades pour être significatif
            wr = (s['winning'] / total * 100) if total > 0 else 0
            pf = s['profit'] / s['loss'] if s['loss'] > 0 else 0
            net = s['profit'] - s['loss']
            combo_results.append({
                'session': key[0].value,
                'condition': key[1].value,
                'trend': key[2].value,
                'trades': total,
                'wr': wr,
                'net': net,
                'pf': pf
            })
    
    # Trier par WR décroissant
    combo_results.sort(key=lambda x: x['wr'], reverse=True)
    
    headers = ["Session", "Condition", "Tendance", "Trades", "WR (%)", "Net (USD)", "Profit Factor"]
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws_combos.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 6
    for combo in combo_results[:15]:  # Top 15 combinaisons
        ws_combos.cell(row=row, column=1).value = combo['session']
        ws_combos.cell(row=row, column=2).value = combo['condition']
        ws_combos.cell(row=row, column=3).value = combo['trend']
        ws_combos.cell(row=row, column=4).value = combo['trades']
        ws_combos.cell(row=row, column=5).value = f"{combo['wr']:.2f}%"
        ws_combos.cell(row=row, column=5).font = Font(bold=True, color="00AA00" if combo['wr'] >= 55 else ("FFA500" if combo['wr'] >= 50 else "AA0000"))
        ws_combos.cell(row=row, column=6).value = f"{combo['net']:.2f}"
        ws_combos.cell(row=row, column=6).font = Font(color="00AA00" if combo['net'] > 0 else "AA0000")
        ws_combos.cell(row=row, column=7).value = f"{combo['pf']:.2f}"
        
        for col in range(1, len(headers) + 1):
            ws_combos.cell(row=row, column=col).border = border
        row += 1
    
    for col in range(1, len(headers) + 1):
        ws_combos.column_dimensions[get_column_letter(col)].width = 18
    
    # Sauvegarder
    wb.save(output_file)
    print(f"OK Rapport Excel genere: {output_file}")


def run_backtest_engine(bot, config, symbol_stats):
    """
    Lance la boucle de backtest (même stratégie que la prod).
    Utilisé par run_backtest.py et run_backtest_last_7_days.py.
    Modifie bot.closed_trades, bot.open_trades, bot.closed_trades_3r, etc.
    """
    use_daily_preferred = config.get('use_daily_preferred_symbol', True)
    one_symbol_at_a_time = config.get('one_symbol_at_a_time', True)
    preferred_by_day = config.get('preferred_symbol_by_day') or {}
    use_next_bar_open_for_entry = config.get('use_next_bar_open_for_entry', True)

    events = []
    for sym in bot.symbols:
        if sym not in bot.historical_data:
            continue
        d = bot.historical_data[sym]
        for i in range(50, len(d)):
            ts = d.index[i]
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            events.append((ts, sym, i))
    events.sort(key=lambda x: x[0])
    total_bars = len(events)
    processed_bars = 0
    bars_skipped_daily_limit = 0
    if events:
        t0, t1 = events[0][0], events[-1][0]
        print(f"   Timeline réelle: {t0.strftime('%Y-%m-%d %H:%M')} → {t1.strftime('%Y-%m-%d %H:%M')} ({total_bars} barres)")

    for current_bar_time, symbol, bar_index in events:
        df = bot.historical_data[symbol]
        # Données jusqu'à bar_index inclus : market_data.iloc[-1] = barre courante (fermée), comme en prod (barres fermées)
        data_index = bar_index
        market_data = bot.get_market_data_at_index(symbol, data_index)
        if market_data is None:
            continue
        current_bar = df.iloc[bar_index]

        if symbol in bot.open_trades and len(bot.open_trades[symbol]) > 0:
            trades_to_close = []
            for trade_index, trade in enumerate(bot.open_trades[symbol]):
                should_close = False
                if trade.type == TradeType.LONG:
                    if current_bar['low'] <= trade.stop_loss:
                        trade.exit_price = trade.stop_loss
                        trade.exit_time = df.index[bar_index]
                        trade.exit_bar_index = bar_index
                        trade.exit_reason = "SL"
                        profit = bot.calculate_profit(symbol, trade.entry_price, trade.exit_price, trade.lot_size, TradeType.LONG)
                        trade.profit = profit
                        bot.current_balance += profit
                        bot.equity = bot.current_balance
                        if profit < 0:
                            bot.last_loss_time = df.index[bar_index]
                        bot.closed_trades.append(trade)
                        should_close = True
                    elif current_bar['high'] >= trade.take_profit:
                        trade.exit_price = trade.take_profit
                        trade.exit_time = df.index[bar_index]
                        trade.exit_bar_index = bar_index
                        trade.exit_reason = "TP"
                        profit = bot.calculate_profit(symbol, trade.entry_price, trade.exit_price, trade.lot_size, TradeType.LONG)
                        trade.profit = profit
                        bot.current_balance += profit
                        bot.equity = bot.current_balance
                        bot.closed_trades.append(trade)
                        should_close = True
                elif trade.type == TradeType.SHORT:
                    if current_bar['high'] >= trade.stop_loss:
                        trade.exit_price = trade.stop_loss
                        trade.exit_time = df.index[bar_index]
                        trade.exit_bar_index = bar_index
                        trade.exit_reason = "SL"
                        profit = bot.calculate_profit(symbol, trade.entry_price, trade.exit_price, trade.lot_size, TradeType.SHORT)
                        trade.profit = profit
                        bot.current_balance += profit
                        bot.equity = bot.current_balance
                        if profit < 0:
                            bot.last_loss_time = df.index[bar_index]
                        bot.closed_trades.append(trade)
                        should_close = True
                    elif current_bar['low'] <= trade.take_profit:
                        trade.exit_price = trade.take_profit
                        trade.exit_time = df.index[bar_index]
                        trade.exit_bar_index = bar_index
                        trade.exit_reason = "TP"
                        profit = bot.calculate_profit(symbol, trade.entry_price, trade.exit_price, trade.lot_size, TradeType.SHORT)
                        trade.profit = profit
                        bot.current_balance += profit
                        bot.equity = bot.current_balance
                        bot.closed_trades.append(trade)
                        should_close = True
                if should_close:
                    trades_to_close.append(trade_index)
            for trade_index in reversed(trades_to_close):
                bot.open_trades[symbol].pop(trade_index)

        if symbol in bot.open_trades_3r and len(bot.open_trades_3r[symbol]) > 0:
            trades_to_close_3r = []
            for trade_index, trade_3r in enumerate(bot.open_trades_3r[symbol]):
                should_close = False
                if trade_3r.type == TradeType.LONG:
                    if current_bar['low'] <= trade_3r.stop_loss:
                        trade_3r.exit_price = trade_3r.stop_loss
                        trade_3r.exit_time = df.index[bar_index]
                        trade_3r.exit_bar_index = bar_index
                        trade_3r.exit_reason = "SL"
                        profit = bot.calculate_profit(symbol, trade_3r.entry_price, trade_3r.exit_price, trade_3r.lot_size, TradeType.LONG)
                        trade_3r.profit = profit
                        bot.closed_trades_3r.append(trade_3r)
                        should_close = True
                    elif current_bar['high'] >= trade_3r.take_profit:
                        trade_3r.exit_price = trade_3r.take_profit
                        trade_3r.exit_time = df.index[bar_index]
                        trade_3r.exit_bar_index = bar_index
                        trade_3r.exit_reason = "TP"
                        profit = bot.calculate_profit(symbol, trade_3r.entry_price, trade_3r.exit_price, trade_3r.lot_size, TradeType.LONG)
                        trade_3r.profit = profit
                        bot.closed_trades_3r.append(trade_3r)
                        should_close = True
                elif trade_3r.type == TradeType.SHORT:
                    if current_bar['high'] >= trade_3r.stop_loss:
                        trade_3r.exit_price = trade_3r.stop_loss
                        trade_3r.exit_time = df.index[bar_index]
                        trade_3r.exit_bar_index = bar_index
                        trade_3r.exit_reason = "SL"
                        profit = bot.calculate_profit(symbol, trade_3r.entry_price, trade_3r.exit_price, trade_3r.lot_size, TradeType.SHORT)
                        trade_3r.profit = profit
                        bot.closed_trades_3r.append(trade_3r)
                        should_close = True
                    elif current_bar['low'] <= trade_3r.take_profit:
                        trade_3r.exit_price = trade_3r.take_profit
                        trade_3r.exit_time = df.index[bar_index]
                        trade_3r.exit_bar_index = bar_index
                        trade_3r.exit_reason = "TP"
                        profit = bot.calculate_profit(symbol, trade_3r.entry_price, trade_3r.exit_price, trade_3r.lot_size, TradeType.SHORT)
                        trade_3r.profit = profit
                        bot.closed_trades_3r.append(trade_3r)
                        should_close = True
                if should_close:
                    trades_to_close_3r.append(trade_index)
            for trade_index in reversed(trades_to_close_3r):
                bot.open_trades_3r[symbol].pop(trade_index)
            if symbol in bot.open_trades_3r and len(bot.open_trades_3r[symbol]) == 0:
                del bot.open_trades_3r[symbol]

        if symbol in bot.open_trades and len(bot.open_trades[symbol]) == 0:
            del bot.open_trades[symbol]

        current_bar_time = df.index[bar_index]
        current_date_obj = current_bar_time.date()
        if config.get('use_daily_loss_in_backtest', False):
            can_trade, _ = bot.can_trade_today(current_date_obj)
            if not can_trade:
                bars_skipped_daily_limit += 1
                continue
        if symbol not in symbol_stats:
            symbol_stats[symbol] = {'signals_detected': 0, 'trades_opened': 0, 'signals_blocked': 0}
        if symbol in bot.last_bar_time and current_bar_time <= bot.last_bar_time[symbol]:
            continue
        bot.last_bar_time[symbol] = current_bar_time

        if use_daily_preferred and preferred_by_day:
            weekday = current_bar_time.weekday() if hasattr(current_bar_time, 'weekday') else getattr(current_bar_time, 'to_pydatetime', lambda: current_bar_time)().weekday()
            preferred_symbol = preferred_by_day.get(weekday)
            if preferred_symbol is not None and symbol != preferred_symbol:
                continue

        def has_other_symbol_with_positions():
            if not one_symbol_at_a_time:
                return False
            for s, trades_list in bot.open_trades.items():
                if s != symbol and trades_list:
                    return True
            return False

        if ALLOW_LONG:
            long_signal = bot.check_long_entry(market_data, symbol, current_bar_time)
            if long_signal:
                symbol_stats[symbol]['signals_detected'] = symbol_stats[symbol].get('signals_detected', 0) + 1
                if not False:
                    # Prix d'entrée : open barre suivante (réaliste, proche du fill réel) ou close barre courante
                    if use_next_bar_open_for_entry and bar_index + 1 < len(df):
                        entry_price = float(df.iloc[bar_index + 1]['open'])
                    else:
                        entry_price = current_bar['close']
                    stop_loss = bot.find_last_low(symbol, market_data, 10)
                    stop_distance = entry_price - stop_loss
                    sl_distance_pct = abs(entry_price - stop_loss) / entry_price if entry_price > 0 else 0
                    rr_ratio = bot.get_risk_reward_ratio(market_data)
                    take_profit = entry_price + (stop_distance * rr_ratio)
                    lot_size = bot.calculate_lot_size(symbol, entry_price, stop_loss)
                    if lot_size <= 0:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif stop_loss <= 0 or stop_loss >= entry_price:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif sl_distance_pct > 0.05:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif has_other_symbol_with_positions():
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    else:
                        symbol_stats[symbol]['trades_opened'] = symbol_stats[symbol].get('trades_opened', 0) + 1
                        trade = SimulatedTrade(
                            symbol=symbol, type=TradeType.LONG,
                            entry_price=entry_price, stop_loss=stop_loss, take_profit=take_profit, lot_size=lot_size,
                            entry_time=current_bar_time, entry_bar_index=bar_index, risk_reward_ratio=rr_ratio
                        )
                        bot.classify_trade(trade, market_data)
                        if symbol not in bot.open_trades:
                            bot.open_trades[symbol] = []
                        bot.open_trades[symbol].append(trade)
                        bot.record_trade(symbol, TradeType.LONG, current_bar_time)
                        if abs(rr_ratio - 1.5) < 0.01:
                            take_profit_3r = entry_price + (stop_distance * 3.0)
                            trade_3r = SimulatedTrade(
                                symbol=symbol, type=TradeType.LONG,
                                entry_price=entry_price, stop_loss=stop_loss, take_profit=take_profit_3r, lot_size=lot_size,
                                entry_time=current_bar_time, entry_bar_index=bar_index, risk_reward_ratio=3.0
                            )
                            bot.classify_trade(trade_3r, market_data)
                            if symbol not in bot.open_trades_3r:
                                bot.open_trades_3r[symbol] = []
                            bot.open_trades_3r[symbol].append(trade_3r)

        if ALLOW_SHORT:
            short_signal = bot.check_short_entry(market_data, symbol, current_bar_time)
            if short_signal:
                symbol_stats[symbol]['signals_detected'] = symbol_stats[symbol].get('signals_detected', 0) + 1
                if not False:
                    # Prix d'entrée : open barre suivante (réaliste) ou close barre courante
                    if use_next_bar_open_for_entry and bar_index + 1 < len(df):
                        entry_price = float(df.iloc[bar_index + 1]['open'])
                    else:
                        entry_price = current_bar['close']
                    stop_loss = bot.find_last_high(symbol, market_data, 10)
                    stop_distance = stop_loss - entry_price
                    sl_distance_pct = abs(stop_loss - entry_price) / entry_price if entry_price > 0 else 0
                    rr_ratio = bot.get_risk_reward_ratio(market_data)
                    take_profit = entry_price - (stop_distance * rr_ratio)
                    lot_size = bot.calculate_lot_size(symbol, entry_price, stop_loss)
                    if lot_size <= 0:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif stop_loss <= 0 or stop_loss <= entry_price:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif sl_distance_pct > 0.05:
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    elif has_other_symbol_with_positions():
                        symbol_stats[symbol]['signals_blocked'] = symbol_stats[symbol].get('signals_blocked', 0) + 1
                    else:
                        symbol_stats[symbol]['trades_opened'] = symbol_stats[symbol].get('trades_opened', 0) + 1
                        trade = SimulatedTrade(
                            symbol=symbol, type=TradeType.SHORT,
                            entry_price=entry_price, stop_loss=stop_loss, take_profit=take_profit, lot_size=lot_size,
                            entry_time=current_bar_time, entry_bar_index=bar_index, risk_reward_ratio=rr_ratio
                        )
                        bot.classify_trade(trade, market_data)
                        if symbol not in bot.open_trades:
                            bot.open_trades[symbol] = []
                        bot.open_trades[symbol].append(trade)
                        bot.record_trade(symbol, TradeType.SHORT, current_bar_time)
                        if abs(rr_ratio - 1.5) < 0.01:
                            take_profit_3r = entry_price - (stop_distance * 3.0)
                            trade_3r = SimulatedTrade(
                                symbol=symbol, type=TradeType.SHORT,
                                entry_price=entry_price, stop_loss=stop_loss, take_profit=take_profit_3r, lot_size=lot_size,
                                entry_time=current_bar_time, entry_bar_index=bar_index, risk_reward_ratio=3.0
                            )
                            bot.classify_trade(trade_3r, market_data)
                            if symbol not in bot.open_trades_3r:
                                bot.open_trades_3r[symbol] = []
                            bot.open_trades_3r[symbol].append(trade_3r)

        bot.equity_curve.append(bot.equity)
        processed_bars += 1
        if processed_bars % 10000 == 0:
            print(f"   Progression: {processed_bars}/{total_bars} bougies traitées...")

    if bars_skipped_daily_limit > 0:
        print(f"\n   ⚠️  Bougies ignorées (limite perte quotidienne): {bars_skipped_daily_limit}")
    for symbol in bot.symbols:
        stats_s = symbol_stats.get(symbol, {})
        signals = stats_s.get('signals_detected', 0)
        trades = stats_s.get('trades_opened', 0)
        blocked = stats_s.get('signals_blocked', 0)
        closed_trades = len([t for t in bot.closed_trades if t.symbol == symbol])
        print(f"   {symbol}: signaux={signals}, bloqués={blocked}, ouverts={trades}, fermés={closed_trades}")

    for symbol, trades_list in list(bot.open_trades.items()):
        df = bot.historical_data[symbol]
        last_bar = df.iloc[-1]
        for trade in trades_list:
            trade.exit_price = last_bar['close']
            trade.exit_time = df.index[-1]
            trade.exit_bar_index = len(df) - 1
            trade.exit_reason = "END"
            profit = bot.calculate_profit(symbol, trade.entry_price, trade.exit_price, trade.lot_size, trade.type)
            trade.profit = profit
            bot.current_balance += profit
            bot.closed_trades.append(trade)
        del bot.open_trades[symbol]

    for symbol, trades_list_3r in list(bot.open_trades_3r.items()):
        df = bot.historical_data[symbol]
        last_bar = df.iloc[-1]
        for trade_3r in trades_list_3r:
            trade_3r.exit_price = last_bar['close']
            trade_3r.exit_time = df.index[-1]
            trade_3r.exit_bar_index = len(df) - 1
            trade_3r.exit_reason = "END"
            profit = bot.calculate_profit(symbol, trade_3r.entry_price, trade_3r.exit_price, trade_3r.lot_size, trade_3r.type)
            trade_3r.profit = profit
            bot.closed_trades_3r.append(trade_3r)
        del bot.open_trades_3r[symbol]


def main():
    """Point d'entrée principal"""
    print("=" * 70)
    print("BACKTEST - Strategie EMA 20 / SMA 50 (Croisement)")
    print("   Timeframe M5 | LONG et SHORT activés")
    print("=" * 70)
    
    # Charger la configuration
    config = load_config()
    
    # Créer le bot de backtest
    bot = MT5BacktestBot(
        login=config['login'],
        password=config['password'],
        server=config['server'],
        symbols=config['symbols'],
        risk_percent=config['risk'],
        max_daily_loss=config['max_daily_loss'],
        initial_balance=config['initial_balance']
    )
    
    # Charger les données historiques pour chaque symbole
    use_all_available = getattr(config, 'USE_ALL_AVAILABLE_DATA', True)
    months_back = config.get('months_back') or 0
    
    print("\nChargement des donnees historiques M5...")
    if months_back > 0:
        print(f"   Mode: Derniers {months_back} mois (plage de dates, pas de limite 50k barres)")
    elif use_all_available:
        print("   Mode: Récupération de TOUTES les données disponibles (maximum possible)")
    else:
        print(f"   Mode: Récupération de {config['years_back']} ans de données")
    
    validated_symbols = []
    symbol_stats = {}  # Statistiques par symbole pour diagnostic
    for symbol in config['symbols']:
        print(f"\n{'='*70}")
        print(f"📊 CHARGEMENT DES DONNÉES POUR {symbol}")
        print(f"{'='*70}")
        df = bot.load_historical_data(
            symbol,
            years=config['years_back'],
            use_all_available=use_all_available if months_back <= 0 else False,
            last_n_months=months_back if months_back > 0 else None
        )
        if df is not None and len(df) > 0:
            # Utiliser le symbole validé (peut être différent de celui dans config)
            validated_symbol = df.index.name if hasattr(df.index, 'name') else symbol
            # Stocker avec le symbole original comme clé
            bot.historical_data[symbol] = df
            validated_symbols.append(symbol)
            period_days = (df.index[-1] - df.index[0]).days if len(df) > 1 else 0
            symbol_stats[symbol] = {
                'bars_loaded': len(df),
                'period_start': df.index[0],
                'period_end': df.index[-1],
                'period_days': period_days,
                'signals_detected': 0,
                'trades_opened': 0,
                'signals_blocked': 0
            }
            print(f"✅ {symbol}: {len(df)} bougies chargées")
            print(f"   Période: {df.index[0]} → {df.index[-1]} ({period_days} jours)")
            if months_back > 0 and period_days < (months_back * 25):
                print(f"   ⚠️ Données courtes: {period_days}j au lieu de ~{months_back*30}j. Téléchargez l'historique dans MT5 (graphique M5, défilez jusqu'à aujourd'hui).")
        else:
            print(f"❌ ATTENTION: Impossible de charger les donnees pour {symbol}")
            symbol_stats[symbol] = {'bars_loaded': 0, 'error': 'Données non chargées'}
    
    # Charger les données H1 pour l'analyse de tendance supérieure
    if USE_H1_TREND_FILTER:
        print("\nChargement des donnees historiques H1 (analyse de tendance)...")
        for symbol in validated_symbols:
            df_h1 = bot.load_h1_data(
                symbol,
                years=config['years_back'],
                use_all_available=config.get('use_all_available', True) if months_back <= 0 else False,
                last_n_months=months_back if months_back > 0 else None
            )
            if df_h1 is not None:
                bot.h1_data[symbol] = df_h1
                print(f"   OK {symbol}: {len(df_h1)} bougies H1 chargees")
            else:
                print(f"   ATTENTION: Impossible de charger les donnees H1 pour {symbol}")
    
    # Mettre à jour les symboles avec ceux validés
    if validated_symbols:
        bot.symbols = validated_symbols
        print(f"\n✅ Symboles valides pour le backtest: {', '.join(validated_symbols)}")
        print(f"\n📊 RÉSUMÉ DES DONNÉES CHARGÉES:")
        for symbol, stats in symbol_stats.items():
            if stats.get('bars_loaded', 0) > 0:
                print(f"   {symbol}: {stats['bars_loaded']} bougies ({stats['period_start']} → {stats['period_end']})")
            else:
                print(f"   {symbol}: ❌ Aucune donnée chargée")
    else:
        print("❌ ERREUR: Aucun symbole valide trouve. Impossible de lancer le backtest.")
        sys.exit(1)
    
    if not bot.historical_data:
        print("❌ ERREUR: Aucune donnee historique chargee. Arret.")
        sys.exit(1)
    
    # Quand months_back > 0: forcer la fenêtre exacte (début = fin - N mois) pour garantir la période demandée
    if months_back > 0:
        # Utiliser explicitement months_back pour la durée (pas years_back)
        n_days = int(months_back * 30)
        end_ts = max(pd.Timestamp(df.index[-1]) for df in bot.historical_data.values())
        start_ts = end_ts - pd.Timedelta(days=n_days)
        # Aligner le timezone si l'index des données est timezone-aware (évite de perdre des barres)
        sample_df = next(iter(bot.historical_data.values()))
        if hasattr(sample_df.index, 'tz') and sample_df.index.tz is not None:
            tz = sample_df.index.tz
            start_ts = start_ts.tz_localize(tz) if start_ts.tzinfo is None else start_ts.tz_convert(tz)
            end_ts = end_ts.tz_localize(tz) if end_ts.tzinfo is None else end_ts.tz_convert(tz)
        print(f"\n📅 Fenêtre forcée: {months_back} derniers mois ({start_ts.strftime('%Y-%m-%d')} → {end_ts.strftime('%Y-%m-%d')})")
        for symbol in list(bot.historical_data.keys()):
            df = bot.historical_data[symbol]
            before = len(df)
            df = df[(df.index >= start_ts) & (df.index <= end_ts)].copy()
            df = df[~df.index.duplicated(keep='first')]
            df.sort_index(inplace=True)
            if len(df) < 50:
                print(f"   ⚠️ {symbol}: trop peu de barres après filtre ({len(df)}), on garde les données telles quelles")
                continue
            bot.historical_data[symbol] = df
            symbol_stats[symbol]['bars_loaded'] = len(df)
            symbol_stats[symbol]['period_start'] = df.index[0]
            symbol_stats[symbol]['period_end'] = df.index[-1]
            print(f"   {symbol}: {before} → {len(df)} barres ({df.index[0].strftime('%Y-%m-%d')} → {df.index[-1].strftime('%Y-%m-%d')})")
        if bot.h1_data:
            for symbol in list(bot.h1_data.keys()):
                df_h1 = bot.h1_data[symbol]
                df_h1 = df_h1[(df_h1.index >= start_ts) & (df_h1.index <= end_ts)].copy()
                df_h1 = df_h1[~df_h1.index.duplicated(keep='first')].sort_index()
                if len(df_h1) > 0:
                    bot.h1_data[symbol] = df_h1
    
    # Lancer le backtest
    print("\nDemarrage du backtest...")
    print("   Cela peut prendre plusieurs minutes...\n")
    
    # Trouver la période commune à tous les symboles
    all_dates = []
    for symbol, df in bot.historical_data.items():
        all_dates.extend(df.index.tolist())
    
    if not all_dates:
        print("ERREUR: Aucune date trouvee")
        sys.exit(1)
    
    # Déterminer la période de backtest
    min_date = max(df.index[0] for df in bot.historical_data.values())
    max_date_actual = min(df.index[-1] for df in bot.historical_data.values())  # fin réelle (limitée par l'actif qui s'arrête le plus tôt)
    max_date_latest = max(df.index[-1] for df in bot.historical_data.values())
    
    print(f"Periode de backtest: {min_date} a {max_date_actual}")
    if max_date_actual != max_date_latest:
        # Un ou plusieurs actifs ont moins d'historique → on indique lesquels
        for sym, df in bot.historical_data.items():
            if df.index[-1] < max_date_latest:
                print(f"   ⚠️ {sym} s'arrête le {df.index[-1].strftime('%Y-%m-%d')} (données manquantes après cette date)")
        print(f"   → Pour avoir les 8 derniers mois complets, téléchargez l'historique MT5 jusqu'à aujourd'hui pour tous les actifs.")
    
    print(f"\n{'='*70}")
    print(f"🔄 DÉMARRAGE DU BACKTEST (timeline unifiée)")
    print(f"{'='*70}")
    use_daily_preferred = config.get('use_daily_preferred_symbol', True)
    one_symbol_at_a_time = config.get('one_symbol_at_a_time', True)
    preferred_by_day = config.get('preferred_symbol_by_day') or {}
    total_bars = sum(max(0, len(bot.historical_data[s]) - 50) for s in bot.symbols if s in bot.historical_data)
    print(f"\n📊 {len(bot.symbols)} actif(s) à comparer: {', '.join(bot.symbols)}")
    print(f"   Timeline: {total_bars} bougies au total (ordre chronologique)")
    if use_daily_preferred and preferred_by_day:
        day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
        parts = [f"{day_names[k]}={v}" for k, v in sorted(preferred_by_day.items()) if 0 <= k < 5]
        if parts:
            print(f"   📅 Actif du jour (config): " + ", ".join(parts))
    if one_symbol_at_a_time:
        print(f"   🔒 Un seul actif en position à la fois (comme en prod)")
    if len(bot.symbols) > 1:
        print(f"   ⚠️  Limite perte quotidienne ({bot.max_daily_loss:.0f}) partagée entre tous les actifs.")
    
    run_backtest_engine(bot, config, symbol_stats)
    
    # Diagnostic: période réelle des sorties de trades (pour vérifier les 8 mois)
    if bot.closed_trades:
        exit_times = []
        for t in bot.closed_trades:
            ts = t.exit_time
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            exit_times.append(ts)
        min_exit = min(exit_times)
        max_exit = max(exit_times)
        months_count = Counter((t.year, t.month) for t in exit_times)
        print(f"\n📅 Période des sorties de trades: {min_exit.strftime('%Y-%m-%d')} → {max_exit.strftime('%Y-%m-%d')}")
        print(f"   Trades par mois: {dict(sorted(months_count.items()))}")
    
    # Calculer et afficher les statistiques
    stats = calculate_stats(bot)
    print_stats(stats)
    
    # Afficher les analytics avancées
    print_advanced_analytics(bot)
    
    # Générer le rapport Excel (avec période complète pour l'onglet Par Mois)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"backtest_report_{timestamp}.xlsx"
    generate_excel_report(bot, stats, excel_file, period_start=min_date, period_end=max_date_actual)
    
    # Analyse détaillée
    print("\n" + "=" * 70)
    print("LANCEMENT DE L'ANALYSE DETAILLEE...")
    print("=" * 70)
    try:
        from analyze_backtest_detailed import generate_detailed_report
        generate_detailed_report(bot, stats)
    except ImportError:
        print("   Script d'analyse détaillée non disponible")
    except Exception as e:
        print(f"   Erreur lors de l'analyse: {e}")
    
    print("\nOK Backtest termine!")
    mt5.shutdown()

if __name__ == "__main__":
    from ema_mt5_bot_backtest import TradeType, SimulatedTrade
    import MetaTrader5 as mt5
    main()
