#!/usr/bin/env python3
"""
backtest_optimizer_wr.py - Round 2: optimisation du Win Rate

Base: H1 2 barres + cooldown 3 (meilleure variante du round 1)
Focus: ameliorer le WR au-dessus de 51% tout en gardant le rendement

Axes testes:
  - Filtre tendance M5 (bougie cloture au-dessus/en-dessous EMA20)
  - Filtre momentum (la bougie va dans le sens du trade)
  - Min distance SL (eviter les SL trop serres qui se font toucher)
  - Filtre ATR min (eviter les marches trop calmes = faux signaux)
  - Sessions restreintes (retirer ASIA si elle baisse le WR)
  - Combinaisons des meilleurs filtres
"""

import sys
import os

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import List, Dict, Optional
import pandas as pd
import numpy as np

_root = os.path.dirname(os.path.abspath(__file__))
_backtest_dir = os.path.join(_root, 'backtest')
for p in [_root, _backtest_dir]:
    if p not in sys.path:
        sys.path.insert(0, p)

import MetaTrader5 as mt5
import strategy_core
from strategy_core import TradeType, TradingSession


# ============================================================================
# REUSE: config, data loading, lot/profit calc from backtest_optimizer.py
# ============================================================================

def load_config():
    try:
        import config
        return {
            'symbols': getattr(config, 'SYMBOLS', ['US30.cash', 'US100.cash', 'US500.cash']),
            'use_daily_preferred_symbol': getattr(config, 'USE_DAILY_PREFERRED_SYMBOL', True),
            'one_symbol_at_a_time': getattr(config, 'ONE_SYMBOL_AT_A_TIME', True),
            'preferred_symbol_by_day': getattr(config, 'PREFERRED_SYMBOL_BY_DAY', None),
            'risk_percent': getattr(config, 'RISK_PERCENT', 1.0),
            'initial_balance': getattr(config, 'INITIAL_BALANCE', 10000.0),
            'login': config.MT5_LOGIN,
            'password': config.MT5_PASSWORD,
            'server': config.MT5_SERVER,
        }
    except Exception as e:
        print(f"ERREUR config: {e}")
        sys.exit(1)


def connect_mt5(cfg):
    if not mt5.initialize():
        for p in [r"C:\Program Files\MetaTrader\terminal64.exe",
                   r"C:\Program Files\MetaTrader 5\terminal64.exe",
                   r"C:\Program Files\MetaTrader 5 Risk Manager\terminal64.exe"]:
            if os.path.exists(p) and mt5.initialize(path=p):
                break
        else:
            print("Impossible d'initialiser MT5.")
            sys.exit(1)
    if not mt5.login(login=cfg['login'], password=cfg['password'], server=cfg['server']):
        print(f"Login echoue: {mt5.last_error()}")
        sys.exit(1)
    info = mt5.account_info()
    print(f"Connecte: {info.login} @ {info.server}")


def load_data(symbols):
    m5_data, h1_data = {}, {}
    for sym in symbols:
        si = mt5.symbol_info(sym)
        if si is None:
            continue
        if not si.visible:
            mt5.symbol_select(sym, True)
        best = None
        for n in [50000, 100000, 200000, 500000, 1000000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_M5, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            strategy_core.compute_indicators(df)
            m5_data[sym] = df
            print(f"  {sym}: {len(df)} M5")
        best = None
        for n in [10000, 50000, 100000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_H1, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            h1_data[sym] = df
    return m5_data, h1_data


def calc_lot(symbol, entry, sl, balance, risk_pct):
    si = mt5.symbol_info(symbol)
    if not si:
        return 0.0
    sd = abs(entry - sl)
    if sd <= 0:
        return 0.0
    risk = balance * (risk_pct / 100.0)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        rpl = (sd / ts) * tv
    elif si.trade_contract_size > 0:
        rpl = (sd * si.trade_contract_size) / entry
    else:
        rpl = sd * si.point
    if rpl <= 0:
        return 0.0
    lot = risk / rpl
    lot = max(si.volume_min, min(lot, si.volume_max))
    if si.volume_step > 0:
        lot = (lot // si.volume_step) * si.volume_step
    return round(lot, 2)


def calc_profit(symbol, entry, exit_p, lot, ttype):
    si = mt5.symbol_info(symbol)
    if not si:
        return 0.0
    diff = (exit_p - entry) if ttype == TradeType.LONG else (entry - exit_p)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        return (diff / ts) * tv * lot
    if si.trade_contract_size > 0:
        return (diff * si.trade_contract_size * lot) / entry
    return diff * si.point * lot


# ============================================================================
# SIMULATION PARAMETRABLE (base = H1 2bars + cooldown 3)
# ============================================================================

@dataclass
class SimTrade:
    symbol: str
    trade_type: TradeType
    entry_time: datetime
    entry_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    exit_time: Optional[datetime] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    profit: float = 0.0


def run_variant(m5_data, h1_data, cfg,
                # Filtres WR
                require_trend_filter=False,    # bougie cloture au-dessus/en-dessous EMA20
                require_momentum=False,        # bougie dans le sens du trade
                min_sl_atr_ratio=0.0,          # SL minimum en multiple d'ATR (ex: 0.5 = au moins 0.5x ATR)
                min_atr_threshold=0.0,         # ATR minimum en % du prix (ex: 0.001 = 0.1%)
                sessions_allowed=None,         # None = toutes, ou set de TradingSession
                require_body_ratio=0.0,        # ratio body/range minimum de la bougie signal (0-1)
                # Base params
                rr_flat=1.0, rr_trending=1.5,
                atr_sl_mult=1.5, cooldown_bars=3):

    symbols = cfg['symbols']
    use_daily_pref = cfg.get('use_daily_preferred_symbol', True)
    one_at_a_time = cfg.get('one_symbol_at_a_time', True)
    pref_by_day = cfg.get('preferred_symbol_by_day') or {}
    risk_pct = cfg.get('risk_percent', 1.0)
    balance = cfg.get('initial_balance', 10000.0)

    orig_rr_f = strategy_core.RISK_REWARD_RATIO_FLAT
    orig_rr_t = strategy_core.RISK_REWARD_RATIO_TRENDING
    orig_atr = strategy_core.ATR_SL_MULTIPLIER
    strategy_core.RISK_REWARD_RATIO_FLAT = rr_flat
    strategy_core.RISK_REWARD_RATIO_TRENDING = rr_trending
    strategy_core.ATR_SL_MULTIPLIER = atr_sl_mult

    events = []
    for sym in symbols:
        if sym not in m5_data:
            continue
        d = m5_data[sym]
        for i in range(50, len(d)):
            ts = d.index[i]
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            events.append((ts, sym, i))
    events.sort(key=lambda x: x[0])

    open_trades: Dict[str, List[SimTrade]] = {}
    closed_trades: List[SimTrade] = []
    last_bar_time: Dict[str, datetime] = {}
    last_loss_time = None

    for current_bar_time, symbol, bar_index in events:
        df = m5_data[symbol]
        market_data = df.iloc[:bar_index + 1]
        current_bar = df.iloc[bar_index]

        # SL/TP management
        if symbol in open_trades and open_trades[symbol]:
            to_close = []
            for idx, trade in enumerate(open_trades[symbol]):
                hit = None
                if trade.trade_type == TradeType.LONG:
                    if current_bar['low'] <= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['high'] >= trade.take_profit:
                        hit = ('TP', trade.take_profit)
                else:
                    if current_bar['high'] >= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['low'] <= trade.take_profit:
                        hit = ('TP', trade.take_profit)
                if hit:
                    trade.exit_reason, trade.exit_price = hit
                    trade.exit_time = df.index[bar_index]
                    trade.profit = calc_profit(symbol, trade.entry_price, trade.exit_price,
                                               trade.lot_size, trade.trade_type)
                    balance += trade.profit
                    if trade.profit < 0:
                        last_loss_time = current_bar_time
                    closed_trades.append(trade)
                    to_close.append(idx)
            for idx in reversed(to_close):
                open_trades[symbol].pop(idx)
            if symbol in open_trades and not open_trades[symbol]:
                del open_trades[symbol]

        if symbol in last_bar_time and current_bar_time <= last_bar_time[symbol]:
            continue
        last_bar_time[symbol] = current_bar_time

        if use_daily_pref and pref_by_day:
            wd = current_bar_time.weekday() if hasattr(current_bar_time, 'weekday') \
                else current_bar_time.to_pydatetime().weekday()
            if pref_by_day.get(wd) is not None and symbol != pref_by_day.get(wd):
                continue

        # Cooldown
        if cooldown_bars > 0 and last_loss_time is not None:
            if (current_bar_time - last_loss_time).total_seconds() / 300 < cooldown_bars:
                continue

        def has_other_open():
            if not one_at_a_time:
                return False
            for s, tl in open_trades.items():
                if s != symbol and tl:
                    return True
            return False

        # H1 data
        df_h1_f = strategy_core.get_h1_data_at_time(
            h1_data.get(symbol, pd.DataFrame()), current_bar_time
        ) if symbol in h1_data else None

        # Session filter custom
        if sessions_allowed is not None:
            session = strategy_core.get_trading_session(current_bar_time)
            if session not in sessions_allowed:
                continue

        for allow, check_fn, ttype, sl_fn, tp_sign in [
            (strategy_core.ALLOW_LONG, strategy_core.check_long_signal,
             TradeType.LONG, strategy_core.calculate_sl_long, 1),
            (strategy_core.ALLOW_SHORT, strategy_core.check_short_signal,
             TradeType.SHORT, strategy_core.calculate_sl_short, -1),
        ]:
            if not allow:
                continue

            signal = check_fn(market_data, df_h1_f, symbol)
            if not signal:
                continue

            # === FILTRES WR SUPPLEMENTAIRES ===

            curr = market_data.iloc[-1]
            prev = market_data.iloc[-2]

            # Filtre trend: bougie cloture du bon cote de EMA20
            if require_trend_filter:
                ema20 = curr[f'EMA_{strategy_core.EMA_FAST}']
                if ttype == TradeType.LONG and curr['close'] <= ema20:
                    continue
                if ttype == TradeType.SHORT and curr['close'] >= ema20:
                    continue

            # Filtre momentum: bougie dans le sens du trade
            if require_momentum:
                if ttype == TradeType.LONG and curr['close'] <= prev['close']:
                    continue
                if ttype == TradeType.SHORT and curr['close'] >= prev['close']:
                    continue

            # Filtre body ratio: taille du corps vs range (evite les dojis)
            if require_body_ratio > 0:
                body = abs(curr['close'] - curr['open'])
                rng = curr['high'] - curr['low']
                if rng > 0 and body / rng < require_body_ratio:
                    continue

            # Filtre ATR minimum (evite marches trop calmes)
            if min_atr_threshold > 0 and 'ATR' in market_data.columns:
                atr = curr.get('ATR', 0)
                if atr > 0 and curr['close'] > 0:
                    atr_pct = atr / curr['close']
                    if atr_pct < min_atr_threshold:
                        continue

            # Calcul entry/SL/TP
            if bar_index + 1 < len(df):
                entry_price = float(df.iloc[bar_index + 1]['open'])
            else:
                entry_price = float(current_bar['close'])

            stop_loss = sl_fn(market_data, 10)

            if ttype == TradeType.LONG:
                stop_dist = entry_price - stop_loss
                if stop_loss <= 0 or stop_loss >= entry_price:
                    continue
            else:
                stop_dist = stop_loss - entry_price
                if stop_loss <= 0 or stop_loss <= entry_price:
                    continue

            sl_pct = abs(entry_price - stop_loss) / entry_price if entry_price > 0 else 0
            if sl_pct > 0.05:
                continue

            # Filtre SL minimum (evite SL trop serre)
            if min_sl_atr_ratio > 0 and 'ATR' in market_data.columns:
                atr = market_data['ATR'].iloc[-1]
                if atr > 0 and abs(stop_dist) < atr * min_sl_atr_ratio:
                    continue

            rr = strategy_core.get_risk_reward_ratio(market_data)
            take_profit = entry_price + (stop_dist * rr * tp_sign)
            lot = calc_lot(symbol, entry_price, stop_loss, balance, risk_pct)

            if lot <= 0 or has_other_open():
                continue

            trade = SimTrade(
                symbol=symbol, trade_type=ttype,
                entry_time=current_bar_time,
                entry_price=entry_price, stop_loss=stop_loss,
                take_profit=take_profit, lot_size=lot,
            )
            open_trades.setdefault(symbol, []).append(trade)

    strategy_core.RISK_REWARD_RATIO_FLAT = orig_rr_f
    strategy_core.RISK_REWARD_RATIO_TRENDING = orig_rr_t
    strategy_core.ATR_SL_MULTIPLIER = orig_atr

    return closed_trades, balance


def compute_stats(trades, initial, final):
    closed = [t for t in trades if t.exit_reason in ('SL', 'TP')]
    total = len(closed)
    if total == 0:
        return {'trades': 0, 'wr': 0, 'net': 0, 'pf': 0, 'rendement': 0, 'avg_win': 0, 'avg_loss': 0, 'max_dd': 0}
    wins = [t for t in closed if t.profit > 0]
    losses = [t for t in closed if t.profit < 0]
    tw = sum(t.profit for t in wins)
    tl = abs(sum(t.profit for t in losses))
    equity, peak, max_dd = initial, initial, 0
    for t in sorted(closed, key=lambda x: x.entry_time):
        equity += t.profit
        if equity > peak:
            peak = equity
        dd = peak - equity
        if dd > max_dd:
            max_dd = dd
    return {
        'trades': total, 'wins': len(wins), 'losses': len(losses),
        'wr': len(wins)/total*100, 'net': final-initial, 'pf': tw/tl if tl > 0 else 999,
        'rendement': (final-initial)/initial*100,
        'avg_win': tw/len(wins) if wins else 0,
        'avg_loss': tl/len(losses) if losses else 0,
        'max_dd': max_dd,
    }


def main():
    cfg = load_config()
    print("="*70)
    print("OPTIMISEUR WR - Round 2 (base: H1 2bars + cooldown 3)")
    print("="*70)
    connect_mt5(cfg)
    print("\nChargement des donnees...")
    m5_data, h1_data = load_data(cfg['symbols'])
    initial = cfg.get('initial_balance', 10000.0)

    EU_US = {TradingSession.EUROPE, TradingSession.US}
    ALL_SESSIONS = None

    variants = [
        # (nom, kwargs)
        ("BASE (H1 2bars + cd3)",                {}),
        ("+ trend filter",                       {"require_trend_filter": True}),
        ("+ momentum filter",                    {"require_momentum": True}),
        ("+ trend + momentum",                   {"require_trend_filter": True, "require_momentum": True}),
        ("+ body ratio > 0.3",                   {"require_body_ratio": 0.3}),
        ("+ body ratio > 0.5",                   {"require_body_ratio": 0.5}),
        ("+ min SL 0.5x ATR",                    {"min_sl_atr_ratio": 0.5}),
        ("+ min SL 0.75x ATR",                   {"min_sl_atr_ratio": 0.75}),
        ("+ ATR min 0.05%",                      {"min_atr_threshold": 0.0005}),
        ("+ ATR min 0.08%",                      {"min_atr_threshold": 0.0008}),
        ("+ sessions EU+US only",                {"sessions_allowed": EU_US}),
        ("+ trend + EU+US",                      {"require_trend_filter": True, "sessions_allowed": EU_US}),
        ("+ momentum + EU+US",                   {"require_momentum": True, "sessions_allowed": EU_US}),
        ("+ trend + momentum + EU+US",           {"require_trend_filter": True, "require_momentum": True, "sessions_allowed": EU_US}),
        ("+ trend + body>0.3",                   {"require_trend_filter": True, "require_body_ratio": 0.3}),
        ("+ momentum + body>0.3",                {"require_momentum": True, "require_body_ratio": 0.3}),
        ("+ trend + minSL 0.5",                  {"require_trend_filter": True, "min_sl_atr_ratio": 0.5}),
        ("+ momentum + minSL 0.5",               {"require_momentum": True, "min_sl_atr_ratio": 0.5}),
        ("+ trend + mom + body>0.3 + EU+US",     {"require_trend_filter": True, "require_momentum": True,
                                                   "require_body_ratio": 0.3, "sessions_allowed": EU_US}),
        ("+ all filters combined",               {"require_trend_filter": True, "require_momentum": True,
                                                   "require_body_ratio": 0.3, "min_sl_atr_ratio": 0.5,
                                                   "sessions_allowed": EU_US}),
        # RR variants on top of best WR filters
        ("+ trend + RR 1:2",                     {"require_trend_filter": True, "rr_trending": 2.0}),
        ("+ momentum + RR 1:2",                  {"require_momentum": True, "rr_trending": 2.0}),
        ("+ trend + EU+US + RR 1:2",             {"require_trend_filter": True, "sessions_allowed": EU_US, "rr_trending": 2.0}),
    ]

    results = []
    print(f"\nTest de {len(variants)} variantes...\n")

    for i, (name, kwargs) in enumerate(variants, 1):
        print(f"  [{i:2d}/{len(variants)}] {name}...", end=" ", flush=True)
        trades, final = run_variant(m5_data, h1_data, cfg, **kwargs)
        stats = compute_stats(trades, initial, final)
        stats['name'] = name
        results.append(stats)
        print(f"{stats['trades']} trades | WR {stats['wr']:.1f}% | PF {stats['pf']:.2f} | "
              f"Rdt {stats['rendement']:.1f}% | DD {stats['max_dd']:.0f}")

    # Tri par WR (puis rendement en cas d'egalite)
    results.sort(key=lambda x: (x['wr'], x['rendement']), reverse=True)

    print(f"\n{'='*120}")
    print(f"{'CLASSEMENT PAR WIN RATE':^120}")
    print(f"{'='*120}")
    print(f"{'#':>3} {'Variante':<45} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8} {'AvgW':>7} {'AvgL':>7}")
    print("-"*120)
    for i, s in enumerate(results, 1):
        marker = " <--" if i == 1 else ""
        print(f"{i:3d} {s['name']:<45} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$ "
              f"{s.get('avg_win',0):6.0f}$ {s.get('avg_loss',0):6.0f}${marker}")

    # Aussi trier par rendement
    results_rdt = sorted(results, key=lambda x: x['rendement'], reverse=True)
    print(f"\n{'='*120}")
    print(f"{'CLASSEMENT PAR RENDEMENT':^120}")
    print(f"{'='*120}")
    print(f"{'#':>3} {'Variante':<45} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8}")
    print("-"*120)
    for i, s in enumerate(results_rdt[:10], 1):
        print(f"{i:3d} {s['name']:<45} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "WR Optimization"
        headers = ["#", "Variante", "Trades", "WR%", "PF", "Net ($)", "Rdt%", "MaxDD ($)", "AvgWin", "AvgLoss"]
        hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        border = Border(*(Side(style='thin'),)*4)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill, cell.font, cell.border = hf, hfont, border
        for r, s in enumerate(results, 2):
            vals = [r-1, s['name'], s['trades'], round(s['wr'],1), round(s['pf'],2),
                    round(s['net'],2), round(s['rendement'],1), round(s['max_dd'],2),
                    round(s.get('avg_win',0),2), round(s.get('avg_loss',0),2)]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = border
            if r == 2:
                for c in range(1, len(headers)+1):
                    ws.cell(row=r, column=c).font = Font(bold=True, color="00AA00")
        ws.column_dimensions['B'].width = 48
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(_root, f"optimizer_wr_{ts}.xlsx")
        wb.save(fname)
        print(f"\nRapport: {fname}")
    except Exception as e:
        print(f"\n(Excel: {e})")

    mt5.shutdown()
    print("\nTermine.")


if __name__ == "__main__":
    main()
