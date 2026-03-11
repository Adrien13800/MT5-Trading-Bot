#!/usr/bin/env python3
"""
backtest_optimizer_r4.py - Round 4: fine-tuning depuis la base RR 1:2 + Time 48

Base: H1 2bars, cooldown 3, RR 1:2 fixe, time exit 48 bars (4h), ATR SL 1.5x

Axes:
  1. Fine-tune R:R (1.5 -> 2.5)
  2. Fine-tune Time Exit (36 -> 72 bars)
  3. Fine-tune ATR SL (1.2x -> 1.7x)
  4. Cooldown (0 -> 5)
  5. H1 filter (2, 3, off)
  6. Confirmation cross (attendre que le cross tienne N barres)
  7. Gap EMA/SMA min (eviter faux croisements)
  8. Combos des meilleurs
"""

import sys
import os

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict, Optional
import pandas as pd
import numpy as np

_root = os.path.dirname(os.path.abspath(__file__))
_backtest_dir = os.path.join(_root, 'backtest')
for p in [_root, _backtest_dir]:
    if p not in sys.path:
        sys.path.insert(0, p)

import MetaTrader5 as mt5
import strategy_core
from strategy_core import TradeType, TradingSession


# ============================================================================
# MT5 / DATA (avec cache symbol_info)
# ============================================================================

_symbol_info_cache = {}

def get_symbol_info(symbol):
    if symbol not in _symbol_info_cache:
        si = mt5.symbol_info(symbol)
        if si:
            _symbol_info_cache[symbol] = si
    return _symbol_info_cache.get(symbol)


def load_config():
    try:
        import config
        return {
            'symbols': getattr(config, 'SYMBOLS', ['US30.cash', 'US100.cash', 'US500.cash']),
            'use_daily_preferred_symbol': getattr(config, 'USE_DAILY_PREFERRED_SYMBOL', True),
            'one_symbol_at_a_time': getattr(config, 'ONE_SYMBOL_AT_A_TIME', True),
            'preferred_symbol_by_day': getattr(config, 'PREFERRED_SYMBOL_BY_DAY', None),
            'risk_percent': getattr(config, 'RISK_PERCENT', 1.0),
            'initial_balance': getattr(config, 'INITIAL_BALANCE', 10000.0),
            'login': config.MT5_LOGIN,
            'password': config.MT5_PASSWORD,
            'server': config.MT5_SERVER,
        }
    except Exception as e:
        print(f"ERREUR config: {e}")
        sys.exit(1)


def connect_mt5(cfg):
    import time as _time
    mt5.shutdown()
    _time.sleep(1)
    paths = [r"C:\Program Files\MetaTrader\terminal64.exe",
             r"C:\Program Files\MetaTrader 5\terminal64.exe",
             r"C:\Program Files\MetaTrader 5 Risk Manager\terminal64.exe"]
    ok = mt5.initialize()
    if not ok:
        for p in paths:
            if os.path.exists(p):
                ok = mt5.initialize(path=p)
                if ok:
                    break
    if not ok:
        for p in paths:
            if os.path.exists(p):
                os.startfile(p)
                print(f"Lancement de MT5: {p}")
                for _ in range(30):
                    _time.sleep(2)
                    if mt5.initialize(path=p):
                        ok = True
                        break
                if ok:
                    break
    if not ok:
        print(f"Impossible d'initialiser MT5. Erreur: {mt5.last_error()}")
        sys.exit(1)
    if not mt5.login(login=cfg['login'], password=cfg['password'], server=cfg['server']):
        print(f"Login echoue: {mt5.last_error()}")
        sys.exit(1)
    info = mt5.account_info()
    print(f"Connecte: {info.login} @ {info.server}")


def load_data(symbols):
    m5_data, h1_data = {}, {}
    for sym in symbols:
        si = mt5.symbol_info(sym)
        if si is None:
            continue
        if not si.visible:
            mt5.symbol_select(sym, True)
        _symbol_info_cache[sym] = si
        best = None
        for n in [50000, 100000, 200000, 500000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_M5, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            strategy_core.compute_indicators(df)
            m5_data[sym] = df
            print(f"  {sym}: {len(df)} M5")
        best = None
        for n in [10000, 50000, 100000]:
            rates = mt5.copy_rates_from_pos(sym, mt5.TIMEFRAME_H1, 0, n)
            if rates is not None and len(rates) > 0:
                if best is None or len(rates) > len(best):
                    best = rates
                if len(rates) < n:
                    break
        if best is not None:
            df = pd.DataFrame(best)
            df['time'] = pd.to_datetime(df['time'], unit='s')
            df.set_index('time', inplace=True)
            df.sort_index(inplace=True)
            h1_data[sym] = df
    return m5_data, h1_data


def calc_lot(symbol, entry, sl, balance, risk_pct):
    si = get_symbol_info(symbol)
    if not si:
        return 0.0
    sd = abs(entry - sl)
    if sd <= 0:
        return 0.0
    risk = balance * (risk_pct / 100.0)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        rpl = (sd / ts) * tv
    elif si.trade_contract_size > 0:
        rpl = (sd * si.trade_contract_size) / entry
    else:
        rpl = sd * si.point
    if rpl <= 0:
        return 0.0
    lot = risk / rpl
    lot = max(si.volume_min, min(lot, si.volume_max))
    if si.volume_step > 0:
        lot = (lot // si.volume_step) * si.volume_step
    return round(lot, 2)


def calc_profit(symbol, entry, exit_p, lot, ttype):
    si = get_symbol_info(symbol)
    if not si:
        return 0.0
    diff = (exit_p - entry) if ttype == TradeType.LONG else (entry - exit_p)
    tv, ts = si.trade_tick_value, si.trade_tick_size
    if ts > 0 and tv > 0:
        return (diff / ts) * tv * lot
    if si.trade_contract_size > 0:
        return (diff * si.trade_contract_size * lot) / entry
    return diff * si.point * lot


# ============================================================================
# SIMULATION R4
# ============================================================================

@dataclass
class SimTrade:
    symbol: str
    trade_type: TradeType
    entry_time: datetime
    entry_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    exit_time: Optional[datetime] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    profit: float = 0.0


def check_cross_signal_custom(df_m5, df_h1, symbol, ttype,
                               h1_bars=2, use_h1=True,
                               cross_confirm_bars=0,
                               min_ema_gap_pct=0.0):
    """
    Signal de croisement EMA20/SMA50 avec filtres parametrables.

    cross_confirm_bars: N > 0 = le croisement doit tenir depuis N barres
    min_ema_gap_pct: gap minimum entre EMA20 et SMA50 en % du prix (evite faux cross)
    """
    if len(df_m5) < 5:
        return False

    current_time = df_m5.index[-1]
    if hasattr(current_time, 'to_pydatetime'):
        current_time = current_time.to_pydatetime()

    # Session filter
    if not strategy_core.is_valid_trading_session(current_time):
        return False

    # H1 filter
    if use_h1 and df_h1 is not None:
        # Use custom h1_bars count
        n = h1_bars
        if len(df_h1) < n:
            return False
        prices = df_h1.iloc[-n:]['close'].values
        if n == 2:
            if ttype == TradeType.LONG and not (prices[1] > prices[0]):
                return False
            if ttype == TradeType.SHORT and not (prices[1] < prices[0]):
                return False
        elif n == 3:
            if ttype == TradeType.LONG:
                if prices[-1] < prices[0]:
                    return False
                rises = sum(1 for i in range(1, len(prices)) if prices[i] > prices[i - 1])
                if rises < 2:
                    return False
            else:
                if prices[-1] > prices[0]:
                    return False
                falls = sum(1 for i in range(1, len(prices)) if prices[i] < prices[i - 1])
                if falls < 2:
                    return False

    ema_col = f'EMA_{strategy_core.EMA_FAST}'
    sma_col = f'SMA_{strategy_core.SMA_SLOW}'

    current = df_m5.iloc[-1]
    prev = df_m5.iloc[-2]

    ema_curr = current[ema_col]
    sma_curr = current[sma_col]
    ema_prev = prev[ema_col]
    sma_prev = prev[sma_col]

    # Cross check
    if ttype == TradeType.LONG:
        if not (ema_prev < sma_prev and ema_curr > sma_curr):
            return False
    else:
        if not (ema_prev > sma_prev and ema_curr < sma_curr):
            return False

    # Gap minimum filter
    if min_ema_gap_pct > 0 and sma_curr > 0:
        gap_pct = abs(ema_curr - sma_curr) / sma_curr
        if gap_pct < min_ema_gap_pct:
            return False

    # Cross confirmation: le croisement doit tenir depuis N barres
    if cross_confirm_bars > 0 and len(df_m5) > cross_confirm_bars + 2:
        for lookback in range(2, 2 + cross_confirm_bars):
            bar = df_m5.iloc[-lookback]
            if ttype == TradeType.LONG:
                if bar[ema_col] <= bar[sma_col]:
                    return False
            else:
                if bar[ema_col] >= bar[sma_col]:
                    return False

    return True


def run_variant(m5_data, h1_data, cfg,
                # R:R
                rr_flat=2.0, rr_trending=2.0,
                # SL
                atr_sl_mult=1.5,
                # Time exit
                max_trade_minutes=240,
                # Cooldown
                cooldown_bars=3,
                # H1 filter
                h1_bars=2, use_h1=True,
                # Cross confirmation
                cross_confirm_bars=0,
                # EMA/SMA gap
                min_ema_gap_pct=0.0):

    symbols = cfg['symbols']
    use_daily_pref = cfg.get('use_daily_preferred_symbol', True)
    one_at_a_time = cfg.get('one_symbol_at_a_time', True)
    pref_by_day = cfg.get('preferred_symbol_by_day') or {}
    risk_pct = cfg.get('risk_percent', 1.0)
    balance = cfg.get('initial_balance', 10000.0)

    orig_rr_f = strategy_core.RISK_REWARD_RATIO_FLAT
    orig_rr_t = strategy_core.RISK_REWARD_RATIO_TRENDING
    orig_atr = strategy_core.ATR_SL_MULTIPLIER
    strategy_core.RISK_REWARD_RATIO_FLAT = rr_flat
    strategy_core.RISK_REWARD_RATIO_TRENDING = rr_trending
    strategy_core.ATR_SL_MULTIPLIER = atr_sl_mult

    events = []
    for sym in symbols:
        if sym not in m5_data:
            continue
        d = m5_data[sym]
        for i in range(50, len(d)):
            ts = d.index[i]
            if hasattr(ts, 'to_pydatetime'):
                ts = ts.to_pydatetime()
            events.append((ts, sym, i))
    events.sort(key=lambda x: x[0])

    open_trades: Dict[str, List[SimTrade]] = {}
    closed_trades: List[SimTrade] = []
    last_bar_time: Dict[str, datetime] = {}
    last_loss_time = None

    for current_bar_time, symbol, bar_index in events:
        df = m5_data[symbol]
        market_data = df.iloc[:bar_index + 1]
        current_bar = df.iloc[bar_index]

        # Trade management: SL/TP + Time exit
        if symbol in open_trades and open_trades[symbol]:
            to_close = []
            for idx, trade in enumerate(open_trades[symbol]):
                hit = None
                if trade.trade_type == TradeType.LONG:
                    if current_bar['low'] <= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['high'] >= trade.take_profit:
                        hit = ('TP', trade.take_profit)
                else:
                    if current_bar['high'] >= trade.stop_loss:
                        hit = ('SL', trade.stop_loss)
                    elif current_bar['low'] <= trade.take_profit:
                        hit = ('TP', trade.take_profit)

                # Time exit
                if hit is None and max_trade_minutes > 0:
                    elapsed = (current_bar_time - trade.entry_time).total_seconds() / 60
                    if elapsed >= max_trade_minutes:
                        hit = ('TIME', float(current_bar['close']))

                if hit:
                    trade.exit_reason, trade.exit_price = hit
                    trade.exit_time = df.index[bar_index]
                    trade.profit = calc_profit(symbol, trade.entry_price, trade.exit_price,
                                               trade.lot_size, trade.trade_type)
                    balance += trade.profit
                    if trade.profit < 0:
                        last_loss_time = current_bar_time
                    closed_trades.append(trade)
                    to_close.append(idx)
            for idx in reversed(to_close):
                open_trades[symbol].pop(idx)
            if symbol in open_trades and not open_trades[symbol]:
                del open_trades[symbol]

        # Signal detection
        if symbol in last_bar_time and current_bar_time <= last_bar_time[symbol]:
            continue
        last_bar_time[symbol] = current_bar_time

        if use_daily_pref and pref_by_day:
            wd = current_bar_time.weekday() if hasattr(current_bar_time, 'weekday') \
                else current_bar_time.to_pydatetime().weekday()
            if pref_by_day.get(wd) is not None and symbol != pref_by_day.get(wd):
                continue

        if cooldown_bars > 0 and last_loss_time is not None:
            if (current_bar_time - last_loss_time).total_seconds() / 300 < cooldown_bars:
                continue

        def has_other_open():
            if not one_at_a_time:
                return False
            for s, tl in open_trades.items():
                if s != symbol and tl:
                    return True
            return False

        # H1 data
        df_h1_f = strategy_core.get_h1_data_at_time(
            h1_data.get(symbol, pd.DataFrame()), current_bar_time
        ) if symbol in h1_data else None

        for ttype, sl_fn, tp_sign in [
            (TradeType.LONG, strategy_core.calculate_sl_long, 1),
            (TradeType.SHORT, strategy_core.calculate_sl_short, -1),
        ]:
            signal = check_cross_signal_custom(
                market_data, df_h1_f, symbol, ttype,
                h1_bars=h1_bars, use_h1=use_h1,
                cross_confirm_bars=cross_confirm_bars,
                min_ema_gap_pct=min_ema_gap_pct,
            )
            if not signal:
                continue

            if bar_index + 1 < len(df):
                entry_price = float(df.iloc[bar_index + 1]['open'])
            else:
                entry_price = float(current_bar['close'])

            stop_loss = sl_fn(market_data, 10)

            if ttype == TradeType.LONG:
                stop_dist = entry_price - stop_loss
                if stop_loss <= 0 or stop_loss >= entry_price:
                    continue
            else:
                stop_dist = stop_loss - entry_price
                if stop_loss <= 0 or stop_loss <= entry_price:
                    continue

            sl_pct = abs(entry_price - stop_loss) / entry_price if entry_price > 0 else 0
            if sl_pct > 0.05:
                continue

            rr = strategy_core.get_risk_reward_ratio(market_data)
            take_profit = entry_price + (stop_dist * rr * tp_sign)
            lot = calc_lot(symbol, entry_price, stop_loss, balance, risk_pct)

            if lot <= 0 or has_other_open():
                continue

            trade = SimTrade(
                symbol=symbol, trade_type=ttype,
                entry_time=current_bar_time,
                entry_price=entry_price, stop_loss=stop_loss,
                take_profit=take_profit, lot_size=lot,
            )
            open_trades.setdefault(symbol, []).append(trade)

    strategy_core.RISK_REWARD_RATIO_FLAT = orig_rr_f
    strategy_core.RISK_REWARD_RATIO_TRENDING = orig_rr_t
    strategy_core.ATR_SL_MULTIPLIER = orig_atr

    return closed_trades, balance


def compute_stats(trades, initial, final):
    closed = [t for t in trades if t.exit_reason in ('SL', 'TP', 'TIME')]
    total = len(closed)
    if total == 0:
        return {'trades': 0, 'wr': 0, 'net': 0, 'pf': 0, 'rendement': 0,
                'avg_win': 0, 'avg_loss': 0, 'max_dd': 0, 'time_exits': 0}
    wins = [t for t in closed if t.profit > 0]
    losses = [t for t in closed if t.profit < 0]
    time_exits = [t for t in closed if t.exit_reason == 'TIME']
    tw = sum(t.profit for t in wins)
    tl = abs(sum(t.profit for t in losses))
    equity, peak, max_dd = initial, initial, 0
    for t in sorted(closed, key=lambda x: x.entry_time):
        equity += t.profit
        if equity > peak:
            peak = equity
        dd = peak - equity
        if dd > max_dd:
            max_dd = dd
    return {
        'trades': total, 'wins': len(wins), 'losses': len(losses),
        'wr': len(wins) / total * 100, 'net': final - initial,
        'pf': tw / tl if tl > 0 else 999,
        'rendement': (final - initial) / initial * 100,
        'avg_win': tw / len(wins) if wins else 0,
        'avg_loss': tl / len(losses) if losses else 0,
        'max_dd': max_dd,
        'time_exits': len(time_exits),
    }


def main():
    cfg = load_config()
    print("=" * 70)
    print("OPTIMISEUR R4 - Fine-tuning (base: RR 1:2 + Time 48 bars)")
    print("=" * 70)
    connect_mt5(cfg)
    print("\nChargement des donnees...")
    m5_data, h1_data = load_data(cfg['symbols'])
    initial = cfg.get('initial_balance', 10000.0)

    # Base = RR 1:2, Time 240 min, ATR SL 1.5x, cd 3, H1 2 bars
    BASE = {"rr_flat": 2.0, "rr_trending": 2.0, "max_trade_minutes": 240,
            "atr_sl_mult": 1.5, "cooldown_bars": 3, "h1_bars": 2, "use_h1": True}

    variants = [
        # === REF ===
        ("BASE: RR 1:2 + Time 48b", {**BASE}),

        # === CAT 1: Fine-tune R:R ===
        ("RR 1:1.5",  {**BASE, "rr_flat": 1.5, "rr_trending": 1.5}),
        ("RR 1:1.7",  {**BASE, "rr_flat": 1.7, "rr_trending": 1.7}),
        ("RR 1:1.8",  {**BASE, "rr_flat": 1.8, "rr_trending": 1.8}),
        ("RR 1:2.0 (base)", {**BASE}),
        ("RR 1:2.2",  {**BASE, "rr_flat": 2.2, "rr_trending": 2.2}),
        ("RR 1:2.5",  {**BASE, "rr_flat": 2.5, "rr_trending": 2.5}),
        ("RR 1:3.0",  {**BASE, "rr_flat": 3.0, "rr_trending": 3.0}),

        # === CAT 2: Fine-tune Time Exit ===
        ("Time 30b (2h30)",  {**BASE, "max_trade_minutes": 150}),
        ("Time 36b (3h)",    {**BASE, "max_trade_minutes": 180}),
        ("Time 42b (3h30)",  {**BASE, "max_trade_minutes": 210}),
        ("Time 48b (base)",  {**BASE}),
        ("Time 54b (4h30)",  {**BASE, "max_trade_minutes": 270}),
        ("Time 60b (5h)",    {**BASE, "max_trade_minutes": 300}),
        ("Time 72b (6h)",    {**BASE, "max_trade_minutes": 360}),
        ("Time OFF",         {**BASE, "max_trade_minutes": 0}),

        # === CAT 3: Fine-tune ATR SL ===
        ("SL 1.0x ATR", {**BASE, "atr_sl_mult": 1.0}),
        ("SL 1.2x ATR", {**BASE, "atr_sl_mult": 1.2}),
        ("SL 1.3x ATR", {**BASE, "atr_sl_mult": 1.3}),
        ("SL 1.5x (base)", {**BASE}),
        ("SL 1.7x ATR", {**BASE, "atr_sl_mult": 1.7}),
        ("SL 2.0x ATR", {**BASE, "atr_sl_mult": 2.0}),

        # === CAT 4: Cooldown ===
        ("Cooldown 0 (off)", {**BASE, "cooldown_bars": 0}),
        ("Cooldown 1",       {**BASE, "cooldown_bars": 1}),
        ("Cooldown 2",       {**BASE, "cooldown_bars": 2}),
        ("Cooldown 3 (base)", {**BASE}),
        ("Cooldown 5",       {**BASE, "cooldown_bars": 5}),
        ("Cooldown 8",       {**BASE, "cooldown_bars": 8}),

        # === CAT 5: H1 Filter ===
        ("H1 OFF",    {**BASE, "use_h1": False}),
        ("H1 2 bars (base)", {**BASE}),
        ("H1 3 bars", {**BASE, "h1_bars": 3}),

        # === CAT 6: Cross Confirmation ===
        ("Confirm 1 bar",  {**BASE, "cross_confirm_bars": 1}),
        ("Confirm 2 bars", {**BASE, "cross_confirm_bars": 2}),
        ("Confirm 3 bars", {**BASE, "cross_confirm_bars": 3}),

        # === CAT 7: EMA/SMA Gap Minimum ===
        ("Gap min 0.01%",  {**BASE, "min_ema_gap_pct": 0.0001}),
        ("Gap min 0.02%",  {**BASE, "min_ema_gap_pct": 0.0002}),
        ("Gap min 0.05%",  {**BASE, "min_ema_gap_pct": 0.0005}),
        ("Gap min 0.1%",   {**BASE, "min_ema_gap_pct": 0.001}),

        # === CAT 8: Combos ===
        ("RR 1:1.8 + Time 36b",  {**BASE, "rr_flat": 1.8, "rr_trending": 1.8, "max_trade_minutes": 180}),
        ("RR 1:1.8 + Time 60b",  {**BASE, "rr_flat": 1.8, "rr_trending": 1.8, "max_trade_minutes": 300}),
        ("RR 1:2.5 + Time 60b",  {**BASE, "rr_flat": 2.5, "rr_trending": 2.5, "max_trade_minutes": 300}),
        ("RR 1:2.5 + Time 72b",  {**BASE, "rr_flat": 2.5, "rr_trending": 2.5, "max_trade_minutes": 360}),
        ("RR 1:3 + Time 72b",    {**BASE, "rr_flat": 3.0, "rr_trending": 3.0, "max_trade_minutes": 360}),
        ("SL 1.2x + RR 1:2",     {**BASE, "atr_sl_mult": 1.2}),
        ("SL 1.2x + RR 1:2.5",   {**BASE, "atr_sl_mult": 1.2, "rr_flat": 2.5, "rr_trending": 2.5}),
        ("SL 1.3x + RR 1:2",     {**BASE, "atr_sl_mult": 1.3}),
        ("Cd 0 + RR 1:2",        {**BASE, "cooldown_bars": 0}),
        ("Cd 0 + RR 1:2.5",      {**BASE, "cooldown_bars": 0, "rr_flat": 2.5, "rr_trending": 2.5}),
        ("H1 OFF + RR 1:2",      {**BASE, "use_h1": False}),
        ("H1 OFF + RR 1:2.5",    {**BASE, "use_h1": False, "rr_flat": 2.5, "rr_trending": 2.5}),
        ("Confirm 1 + Gap 0.02%", {**BASE, "cross_confirm_bars": 1, "min_ema_gap_pct": 0.0002}),
        ("RR 1:2 + Cd 0 + SL 1.2x", {**BASE, "cooldown_bars": 0, "atr_sl_mult": 1.2}),
        ("RR 1:2.5 + Cd 0 + Time 60b", {**BASE, "cooldown_bars": 0, "rr_flat": 2.5, "rr_trending": 2.5, "max_trade_minutes": 300}),
        ("RR 1:3 + Cd 0 + Time 72b", {**BASE, "cooldown_bars": 0, "rr_flat": 3.0, "rr_trending": 3.0, "max_trade_minutes": 360}),
    ]

    results = []
    n_variants = len(variants)
    print(f"\nTest de {n_variants} variantes...\n")

    for i, (name, kwargs) in enumerate(variants, 1):
        print(f"  [{i:2d}/{n_variants}] {name}...", end=" ", flush=True)
        trades, final = run_variant(m5_data, h1_data, cfg, **kwargs)
        stats = compute_stats(trades, initial, final)
        stats['name'] = name
        results.append(stats)
        print(f"{stats['trades']} trades | WR {stats['wr']:.1f}% | PF {stats['pf']:.2f} | "
              f"Rdt {stats['rendement']:.1f}% | DD {stats['max_dd']:.0f} | "
              f"TIME {stats['time_exits']}")

    # === CLASSEMENT PAR RENDEMENT ===
    results_rdt = sorted(results, key=lambda x: x['rendement'], reverse=True)

    print(f"\n{'=' * 140}")
    print(f"{'CLASSEMENT PAR RENDEMENT (TOP 25)':^140}")
    print(f"{'=' * 140}")
    print(f"{'#':>3} {'Variante':<40} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Net$':>10} {'Rdt%':>8} {'MaxDD$':>8} {'DD%':>6} {'AvgW':>7} {'AvgL':>7} {'TIME':>5}")
    print("-" * 140)
    for i, s in enumerate(results_rdt[:25], 1):
        dd_pct = s['max_dd'] / initial * 100
        marker = " <-- BEST" if i == 1 else ""
        print(f"{i:3d} {s['name']:<40} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['net']:9.0f}$ {s['rendement']:7.1f}% {s['max_dd']:7.0f}$ {dd_pct:5.1f}% "
              f"{s.get('avg_win', 0):6.0f}$ {s.get('avg_loss', 0):6.0f}$ {s.get('time_exits', 0):5d}{marker}")

    # === CLASSEMENT PAR RATIO Rdt/DD ===
    print(f"\n{'=' * 140}")
    print(f"{'CLASSEMENT PAR RENDEMENT / DRAWDOWN (TOP 15)':^140}")
    print(f"{'=' * 140}")
    print(f"{'#':>3} {'Variante':<40} {'Trades':>6} {'WR%':>6} {'PF':>6} {'Rdt%':>8} {'MaxDD$':>8} {'DD%':>6} {'Rdt/DD':>8}")
    print("-" * 140)
    results_cal = sorted(results, key=lambda x: x['rendement'] / (x['max_dd'] / initial * 100) if x['max_dd'] > 0 else 0, reverse=True)
    for i, s in enumerate(results_cal[:15], 1):
        dd_pct = s['max_dd'] / initial * 100
        ratio = s['rendement'] / dd_pct if dd_pct > 0 else 0
        print(f"{i:3d} {s['name']:<40} {s['trades']:6d} {s['wr']:5.1f}% {s['pf']:6.2f} "
              f"{s['rendement']:7.1f}% {s['max_dd']:7.0f}$ {dd_pct:5.1f}% {ratio:7.2f}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "R4 Optimization"
        headers = ["#", "Variante", "Trades", "WR%", "PF", "Net ($)", "Rdt%",
                    "MaxDD ($)", "DD%", "AvgWin", "AvgLoss", "TimeExits"]
        hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        border = Border(*(Side(style='thin'),) * 4)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill, cell.font, cell.border = hf, hfont, border
        for r, s in enumerate(results_rdt, 2):
            dd_pct = s['max_dd'] / initial * 100
            vals = [r - 1, s['name'], s['trades'], round(s['wr'], 1), round(s['pf'], 2),
                    round(s['net'], 2), round(s['rendement'], 1), round(s['max_dd'], 2),
                    round(dd_pct, 1),
                    round(s.get('avg_win', 0), 2), round(s.get('avg_loss', 0), 2),
                    s.get('time_exits', 0)]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = border
            if r == 2:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).font = Font(bold=True, color="00AA00")
        ws.column_dimensions['B'].width = 44
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(_root, f"optimizer_r4_{ts_str}.xlsx")
        wb.save(fname)
        print(f"\nRapport: {fname}")
    except Exception as e:
        print(f"\n(Excel: {e})")

    mt5.shutdown()
    print("\nTermine.")


if __name__ == "__main__":
    main()
